"""
Generador de Modelo TDABC para CardioCentro Diagnóstico Integral S.A.S.
Modelo profesional de costeo basado en Time-Driven Activity-Based Costing
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, PieChart, Reference
import random
from datetime import datetime
import io

class ModeloTDABC:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remover hoja por defecto
        
        # Colores corporativos (Azul médico y Gris)
        self.COLOR_HEADER = "003366"  # Azul oscuro
        self.COLOR_SUBHEADER = "4472C4"  # Azul medio
        self.COLOR_TEXT = "000000"
        self.COLOR_BG_HEADER = "D9E1F2"  # Azul claro fondo
        self.COLOR_INPUT = "E2EFDA"   # Verde claro para inputs
        self.COLOR_CALCULO = "F2F2F2" # Gris para cálculos
        self.COLOR_RESULTADO = "DDEBF7" # Azul muy claro resultados
        self.FUENTE_BASE = "Arial Narrow" # Fuente solicitada
        
        # Datos base
        self.sedes = ["Sede Norte - Bogotá", "Sede Sur - Medellín", "Sede Centro - Cali"]
        self.aseguradoras = ["SURA", "Sanitas", "Compensar", "Salud Total", "Nueva EPS"]
        
        # Servicios Cardiólogicos
        self.servicios = [
            "Ecocardiograma Transtorácico", "Ecocardiograma Transesofágico",
            "Holter 24 Horas", "Holter 48 Horas", "Prueba de Esfuerzo",
            "Electrocardiograma", "MAPA 24 Horas", "Ecocardiograma Doppler",
            "Ecocardiograma de Estrés", "Tilt Test",
            "Estudio Electrofisiológico", "Cardioversión Eléctrica",
            "Implante Marcapasos", "Cateterismo Cardíaco Derecho",
            "Cateterismo Cardíaco Izquierdo", "Angioplastia Coronaria",
            "Ablación por Radiofrecuencia", "Cierre de CIA",
            "Estudio Hemodinámico Completo", "Biopsia Endomiocárdica"
        ]
        
        self.grupos_ocupacionales = [
            ("Cardiólogo Especialista", 12000000),
            ("Cardiólogo General", 8000000),
            ("Médico General", 5000000),
            ("Enfermero Especializado", 3500000),
            ("Enfermero", 2800000),
            ("Técnico Radiólogo", 2500000),
            ("Auxiliar de Enfermería", 1800000),
        ]
        
        self.inicializar_datos_financieros()
    
    def crear_tabla(self, ws, nombre_tabla, rango_datos):
        """Convierte un rango de celdas en una Tabla de Excel oficial"""
        tab = Table(displayName=nombre_tabla, ref=rango_datos)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        return tab

    def crear_estilo_header(self):
        """Crea el estilo para encabezados"""
        font = Font(name=self.FUENTE_BASE, size=11, bold=True, color="FFFFFF")
        fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        return font, fill, alignment, border
    
    def aplicar_estilo_celda(self, cell, tipo="normal"):
        """Aplica estilo a una celda según su tipo"""
        
        if tipo == "input":
            cell.fill = PatternFill(start_color=self.COLOR_INPUT, end_color=self.COLOR_INPUT, fill_type="solid")
        elif tipo == "calculo":
            cell.fill = PatternFill(start_color=self.COLOR_CALCULO, end_color=self.COLOR_CALCULO, fill_type="solid")
        elif tipo == "resultado":
            cell.fill = PatternFill(start_color=self.COLOR_RESULTADO, end_color=self.COLOR_RESULTADO, fill_type="solid")
    
    def ajustar_columnas(self, ws, columnas_info):
        """Ajusta el ancho de las columnas"""
        for col, width in columnas_info.items():
            ws.column_dimensions[col].width = width
    
    def inicializar_datos_financieros(self):
        """Pre-calcula costos indirectos y tasas para consistencia total."""
        self.presupuesto_indirectos = {
            "7305 - Arrendamiento": 150000000, "7310 - Depreciación Equipos": 85000000,
            "7315 - Mantenimiento": 45000000, "7320 - Servicios Públicos": 35000000,
            "7325 - Seguros": 28000000, "7330 - Aseo y Limpieza": 18000000,
            "7335 - Vigilancia": 22000000, "7340 - Papelería": 5000000,
            "7390 - Otros": 15000000
        }
        self.total_indirectos_global = sum(self.presupuesto_indirectos.values())
        self.indirectos_por_sede = {}
        self.tasas_cif_por_sede = {}
        self.salas_por_sede = {"Sede Norte - Bogotá": 4, "Sede Sur - Medellín": 3, "Sede Centro - Cali": 2} # Nombres ajustados
        total_salas = sum(self.salas_por_sede.values())
        self.capacidad_mensual_minutos = 11040 
        
        for sede, salas in self.salas_por_sede.items():
            factor = salas / total_salas
            costo_sede = self.total_indirectos_global * factor
            self.indirectos_por_sede[sede] = costo_sede
            capacidad_sede = salas * self.capacidad_mensual_minutos
            tasa = costo_sede / capacidad_sede
            self.tasas_cif_por_sede[sede] = tasa

    def crear_hoja_parametros(self):
        """Crea la hoja PARAMETROS con configuración general"""
        ws = self.wb.create_sheet("PARAMETROS")
        
        # Título
        ws['A1'] = "CardioCentro Diagnóstico Integral S.A.S."
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = "PARÁMETROS GENERALES DEL MODELO TDABC"
        ws['A2'].font = Font(name='Calibri', size=12, bold=True)
        ws.merge_cells('A2:D2')
        
        # Fecha de actualización
        ws['A3'] = f"Fecha de actualización: {datetime.now().strftime('%d/%m/%Y')}"
        ws.merge_cells('A3:D3')
        
        # Parámetros operacionales
        row = 5
        ws[f'A{row}'] = "PARÁMETROS OPERACIONALES"
        font, fill, alignment, border = self.crear_estilo_header()
        ws[f'A{row}'].font = font
        ws[f'A{row}'].fill = fill
        ws.merge_cells(f'A{row}:B{row}')
        
        parametros = [
            ("Horas mensuales por colaborador", 184, "horas"),
            ("Días hábiles al mes", 22, "días"),
            ("Porcentaje capacidad práctica", 85, "%"),
            ("Factor prestacional sector salud", 1.4077, "factor"),
            ("Prima de servicios", 8.33, "%"),
            ("Cesantías", 8.33, "%"),
            ("Intereses cesantías", 1.00, "%"),
            ("Vacaciones", 4.17, "%"),
            ("ARL (Riesgo III - CATEROGI)", 2.436, "%"),
            ("Salud EPS (empleador)", 0, "%"),
            ("Pensión (empleador)", 12.5, "%"),
            ("Caja compensación", 4.0, "%"),
            ("ICBF (exonerado sector salud)", 0, "%"),
            ("SENA (exonerado sector salud)", 0, "%"),
            ("TOTAL PRESTACIONES LEGALES", 40.77, "%"),
            ("Inflación anual proyectada", 5.5, "%"),
            ("Tasa cambio USD (referencial)", 4100, "COP"),
        ]
        
        row += 1
        ws[f'A{row}'] = "Parámetro"
        ws[f'B{row}'] = "Valor"
        ws[f'C{row}'] = "Unidad"
        for col in ['A', 'B', 'C']:
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
        
        for param, valor, unidad in parametros:
            row += 1
            ws[f'A{row}'] = param
            ws[f'B{row}'] = valor
            ws[f'C{row}'] = unidad
            self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'B{row}'], "input")
            self.aplicar_estilo_celda(ws[f'C{row}'], "calculo")
        
        
        # Nota explicativa sobre prestaciones sector salud
        row += 2
        ws[f'A{row}'] = "NOTA: Empresas del sector salud en Colombia"
        ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color="2C3E50")
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Las empresas del sector salud están EXONERADAS del pago de parafiscales (ICBF 3% + SENA 2%)."
        ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color="2C3E50")
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Sin embargo, SÍ pagan EPS (8.5%), Pensión (12%), Caja Compensación (4%) y ARL según riesgo."
        ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color="2C3E50")
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:C{row}')
        
        # Parámetros de distribución de costos indirectos
        row += 3
        ws[f'A{row}'] = "DISTRIBUCIÓN DE COSTOS INDIRECTOS"
        font, fill, alignment, border = self.crear_estilo_header()
        ws[f'A{row}'].font = font
        ws[f'A{row}'].fill = fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Concepto"
        ws[f'B{row}'] = "Criterio"
        for col in ['A', 'B']:
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
        
        criterios = [
            ("Servicios generales", "Por minutos consumidos"),
            ("Administración", "Por volumen de producción"),
            ("Infraestructura", "Por área utilizada (m²)"),
        ]
        
        for concepto, criterio in criterios:
            row += 1
            ws[f'A{row}'] = concepto
            ws[f'B{row}'] = criterio
            self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
        
        # Información de sedes
        row += 3
        ws[f'A{row}'] = "SEDES OPERATIVAS"
        font, fill, alignment, border = self.crear_estilo_header()
        ws[f'A{row}'].font = font
        ws[f'A{row}'].fill = fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ["Código", "Nombre Sede", "Ciudad", "Área (m²)"]
        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            ws[f'{col}{row}'] = header
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
        
        sedes_info = [
            ("S01", "Sede Norte - Bogotá", "Bogotá", 450),
            ("S02", "Sede Sur - Medellín", "Medellín", 380),
            ("S03", "Sede Centro - Cali", "Cali", 420),
        ]
        
        for codigo, nombre, ciudad, area in sedes_info:
            row += 1
            ws[f'A{row}'] = codigo
            ws[f'B{row}'] = nombre
            ws[f'C{row}'] = ciudad
            ws[f'D{row}'] = area
            for col in ['A', 'B', 'C']:
                self.aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'D{row}'], "input")
        
        self.ajustar_columnas(ws, {'A': 35, 'B': 30, 'C': 20, 'D': 15})
        
    def crear_hoja_nomina(self):
        """Crea la hoja NOMINA con estructura salarial"""
        ws = self.wb.create_sheet("NOMINA")
        
        # Título
        ws['A1'] = "ESTRUCTURA SALARIAL Y NÓMINA"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:H1')
        
        # Encabezados
        headers = [
            "ID", "Grupo Ocupacional", "Salario Base Mensual", 
            "Factor Prestacional", "Costo Total Mensual",
            "Cantidad Personal", "Costo Total Grupo", "Sede"
        ]
        
        row = 3
        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            ws[f'{col}{row}'] = header
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = alignment
        
        # Datos de nómina
        row = 4
        id_counter = 1
        for grupo, salario in self.grupos_ocupacionales:
            for sede in self.sedes:
                cantidad = random.randint(2, 8)
                
                ws[f'A{row}'] = f"G{id_counter:03d}"
                ws[f'B{row}'] = grupo
                ws[f'C{row}'] = salario
                ws[f'D{row}'] = "=PARAMETROS!$B$10"  # Factor prestacional sector salud (40.77%)
                ws[f'E{row}'] = f"=C{row}*D{row}"  # Costo total mensual
                ws[f'F{row}'] = cantidad
                ws[f'G{row}'] = f"=E{row}*F{row}"  # Costo total grupo
                ws[f'H{row}'] = sede
                
                # Aplicar estilos
                self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'C{row}'], "input")
                self.aplicar_estilo_celda(ws[f'D{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'E{row}'], "resultado")
                self.aplicar_estilo_celda(ws[f'F{row}'], "input")
                self.aplicar_estilo_celda(ws[f'G{row}'], "resultado")
                self.aplicar_estilo_celda(ws[f'H{row}'], "calculo")
                
                # Formato moneda
                ws[f'C{row}'].number_format = '$#,##0'
                ws[f'E{row}'].number_format = '$#,##0'
                ws[f'G{row}'].number_format = '$#,##0'
                
                row += 1
                id_counter += 1
        
        # Totales
        ws[f'B{row}'] = "TOTAL NÓMINA MENSUAL"
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'G{row}'] = f"=SUM(G4:G{row-1})"
        ws[f'G{row}'].font = Font(bold=True)
        ws[f'G{row}'].number_format = '$#,##0'
        font, fill, alignment, border = self.crear_estilo_header()
        ws[f'B{row}'].fill = fill
        ws[f'G{row}'].fill = fill
        ws.merge_cells(f'B{row}:F{row}')
        
        self.ajustar_columnas(ws, {
            'A': 10, 'B': 30, 'C': 20, 'D': 18, 
            'E': 20, 'F': 18, 'G': 20, 'H': 25
        })
    
    def crear_hoja_capacidad(self):
        """Crea la hoja CAPACIDAD con cálculo de capacidad práctica"""
        ws = self.wb.create_sheet("CAPACIDAD")
        
        # Título
        ws['A1'] = "ANÁLISIS DE CAPACIDAD PRÁCTICA"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:G1')
        
        # Explicación de capacidad práctica
        ws['A2'] = "NOTA: La capacidad práctica (85%) es el tiempo REAL disponible. Se descuenta ~15% por pausas, reuniones, capacitación, etc."
        ws['A2'].font = Font(name='Calibri', size=9, italic=True, color="D35400")
        ws['A2'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('A2:G2')
        
        # Encabezados
        headers = [
            "Grupo Ocupacional", "Horas Mensuales", "Minutos Disponibles",
            "% Capacidad Práctica", "Minutos Capacidad Práctica",
            "Personal Disponible", "Total Minutos Grupo"
        ]
        
        row = 4
        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            ws[f'{col}{row}'] = header
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Datos de capacidad
        row = 5  # Iniciar después de los headers (row 4)
        for grupo, _ in self.grupos_ocupacionales:
            ws[f'A{row}'] = grupo
            ws[f'B{row}'] = "=PARAMETROS!$B$7"  # Horas mensuales (184h)
            ws[f'C{row}'] = f"=B{row}*60"  # Minutos disponibles
            ws[f'D{row}'] = "=PARAMETROS!$B$9/100"  # % Capacidad práctica (85%)
            ws[f'E{row}'] = f"=C{row}*D{row}"  # Minutos capacidad práctica
            # Sumar personal de este grupo en todas las sedes
            ws[f'F{row}'] = f'=SUMIF(NOMINA!$B:$B,A{row},NOMINA!$F:$F)'
            ws[f'G{row}'] = f"=E{row}*F{row}"  # Total minutos grupo
            
            # Aplicar estilos
            self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'C{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'D{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'E{row}'], "resultado")
            self.aplicar_estilo_celda(ws[f'F{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'G{row}'], "resultado")
            
            # Formato
            ws[f'D{row}'].number_format = '0%'
            ws[f'E{row}'].number_format = '#,##0'
            ws[f'G{row}'].number_format = '#,##0'
            
            row += 1
        
        # Totales
        ws[f'A{row}'] = "TOTAL CAPACIDAD ORGANIZACIONAL"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'G{row}'] = f"=SUM(G4:G{row-1})"
        ws[f'G{row}'].font = Font(bold=True)
        ws[f'G{row}'].number_format = '#,##0'
        font, fill, alignment, border = self.crear_estilo_header()
        ws[f'A{row}'].fill = fill
        ws[f'G{row}'].fill = fill
        ws.merge_cells(f'A{row}:F{row}')
        
        self.ajustar_columnas(ws, {
            'A': 30, 'B': 18, 'C': 20, 'D': 20,
            'E': 25, 'F': 20, 'G': 25
        })
    
    def crear_hoja_costo_por_minuto(self):
        """Crea la hoja COSTO_POR_MINUTO"""
        ws = self.wb.create_sheet("COSTO_POR_MINUTO")
        
        # Título
        ws['A1'] = "COSTO POR MINUTO - NÚCLEO DEL MODELO TDABC"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:E1')
        
        # Encabezados
        headers = [
            "Grupo Ocupacional", "Costo Total Mensual", 
            "Minutos Capacidad Práctica", "Costo por Minuto", "Sede"
        ]
        
        row = 3
        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            ws[f'{col}{row}'] = header
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Datos
        row = 4
        for grupo, _ in self.grupos_ocupacionales:
            for sede in self.sedes:
                ws[f'A{row}'] = grupo
                # Buscar costo total mensual de este grupo en esta sede
                ws[f'B{row}'] = f'=SUMIFS(NOMINA!$E:$E,NOMINA!$B:$B,A{row},NOMINA!$H:$H,E{row})'
                # Minutos capacidad práctica por persona
                ws[f'C{row}'] = f'=VLOOKUP(A{row},CAPACIDAD!$A:$E,5,FALSE)'
                # Costo por minuto = Costo Total / Minutos Capacidad Práctica
                ws[f'D{row}'] = f'=IF(C{row}>0,B{row}/C{row},0)'
                ws[f'E{row}'] = sede
                
                # Aplicar estilos
                self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'C{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'D{row}'], "resultado")
                self.aplicar_estilo_celda(ws[f'E{row}'], "calculo")
                
                # Formato
                ws[f'B{row}'].number_format = '$#,##0'
                ws[f'C{row}'].number_format = '#,##0'
                ws[f'D{row}'].number_format = '$#,##0.00'
                
                row += 1
        
        self.ajustar_columnas(ws, {
            'A': 30, 'B': 25, 'C': 28, 'D': 20, 'E': 25
        })
    
    def crear_hoja_servicios(self):
        """Crea la hoja SERVICIOS con catálogo de servicios"""
        ws = self.wb.create_sheet("SERVICIOS")
        
        # Título
        ws['A1'] = "CATÁLOGO DE SERVICIOS"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:F1')
        
        # Encabezados
        headers = [
            "Código", "Nombre del Servicio", "Categoría", 
            "Complejidad", "Requiere Insumos", "Estado"
        ]
        
        row = 3
        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            ws[f'{col}{row}'] = header
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Categorías
        categorias = {
            "Diagnóstico No Invasivo": ["Ecocardiograma Transtorácico", "Holter 24 Horas", 
                                         "Holter 48 Horas", "Electrocardiograma", 
                                         "MAPA 24 Horas", "Ecocardiograma Doppler"],
            "Diagnóstico Funcional": ["Prueba de Esfuerzo", "Ecocardiograma de Estrés", 
                                      "Tilt Test"],
            "Diagnóstico Invasivo": ["Ecocardiograma Transesofágico", "Estudio Electrofisiológico", 
                                     "Cateterismo Cardíaco Derecho", "Cateterismo Cardíaco Izquierdo",
                                     "Estudio Hemodinámico Completo", "Biopsia Endomiocárdica"],
            "Terapéutico": ["Cardioversión Eléctrica", "Implante Marcapasos", 
                           "Angioplastia Coronaria", "Ablación por Radiofrecuencia", 
                           "Cierre de CIA"]
        }
        
        # Datos de servicios
        row = 4
        servicio_idx = 1
        for categoria, servicios in categorias.items():
            for servicio in servicios:
                if servicio in self.servicios:
                    # Determinar complejidad
                    if categoria == "Terapéutico":
                        complejidad = "Alta"
                    elif categoria == "Diagnóstico Invasivo":
                        complejidad = "Media-Alta"
                    elif categoria == "Diagnóstico Funcional":
                        complejidad = "Media"
                    else:
                        complejidad = "Baja"
                    
                    ws[f'A{row}'] = f"SV{servicio_idx:03d}"
                    ws[f'B{row}'] = servicio
                    ws[f'C{row}'] = categoria
                    ws[f'D{row}'] = complejidad
                    ws[f'E{row}'] = "Sí" if categoria in ["Diagnóstico Invasivo", "Terapéutico"] else "No"
                    ws[f'F{row}'] = "Activo"
                    
                    # Aplicar estilos
                    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                        self.aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
                    
                    row += 1
                    servicio_idx += 1
        
        self.ajustar_columnas(ws, {
            'A': 12, 'B': 38, 'C': 25, 'D': 15, 'E': 18, 'F': 12
        })
    
    def crear_hoja_ecuaciones_tiempo(self):
        """Crea la hoja ECUACIONES_TIEMPO - núcleo del TDABC"""
        ws = self.wb.create_sheet("ECUACIONES_TIEMPO")
        
        # Título
        ws['A1'] = "ECUACIONES DE TIEMPO TDABC"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = "Núcleo del modelo: Cada servicio consume tiempo de diferentes grupos ocupacionales"
        ws['A2'].font = Font(name='Calibri', size=10, italic=True)
        ws.merge_cells('A2:F2')
        
        ws['A3'] = "FACTOR DE COMPLEJIDAD: Ajusta el tiempo según dificultad del caso. 1.0=normal, 1.2-1.5=complejo, 2.0+=muy complejo. Es AJUSTABLE según su realidad."
        ws['A3'].font = Font(name='Calibri', size=9, italic=True, color="D35400")
        ws['A3'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('A3:F3')
        
        # Encabezados
        headers = [
            "Código Servicio", "Nombre Servicio", "Grupo Ocupacional",
            "Minutos Requeridos", "Factor Complejidad", "Minutos Ajustados"
        ]
        
        row = 5
        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            ws[f'{col}{row}'] = header
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Definición de ecuaciones por servicio
        ecuaciones_servicios = {
            "Ecocardiograma Transtorácico": [
                ("Cardiólogo Especialista", 15, 1.0),
                ("Técnico Radiólogo", 20, 1.0),
                ("Auxiliar de Enfermería", 10, 1.0)
            ],
            "Ecocardiograma Transesofágico": [
                ("Cardiólogo Especialista", 30, 1.2),
                ("Enfermero Especializado", 25, 1.0),
                ("Auxiliar de Enfermería", 15, 1.0)
            ],
            "Holter 24 Horas": [
                ("Cardiólogo General", 20, 1.0),
                ("Técnico Radiólogo", 30, 1.0),
                ("Auxiliar de Enfermería", 15, 1.0)
            ],
            "Holter 48 Horas": [
                ("Cardiólogo General", 25, 1.1),
                ("Técnico Radiólogo", 35, 1.0),
                ("Auxiliar de Enfermería", 20, 1.0)
            ],
            "Prueba de Esfuerzo": [
                ("Cardiólogo General", 30, 1.0),
                ("Enfermero Especializado", 25, 1.0),
                ("Técnico Radiólogo", 15, 1.0)
            ],
            "Electrocardiograma": [
                ("Técnico Radiólogo", 10, 1.0),
                ("Auxiliar de Enfermería", 5, 1.0)
            ],
            "MAPA 24 Horas": [
                ("Cardiólogo General", 20, 1.0),
                ("Enfermero", 25, 1.0),
                ("Auxiliar de Enfermería", 15, 1.0)
            ],
            "Ecocardiograma Doppler": [
                ("Cardiólogo Especialista", 20, 1.0),
                ("Técnico Radiólogo", 25, 1.0),
                ("Auxiliar de Enfermería", 10, 1.0)
            ],
            "Ecocardiograma de Estrés": [
                ("Cardiólogo Especialista", 35, 1.2),
                ("Enfermero Especializado", 30, 1.0),
                ("Técnico Radiólogo", 20, 1.0)
            ],
            "Tilt Test": [
                ("Cardiólogo General", 40, 1.0),
                ("Enfermero Especializado", 45, 1.0),
                ("Auxiliar de Enfermería", 20, 1.0)
            ],
            "Estudio Electrofisiológico": [
                ("Cardiólogo Especialista", 90, 1.5),
                ("Enfermero Especializado", 60, 1.2),
                ("Técnico Radiólogo", 45, 1.0)
            ],
            "Cardioversión Eléctrica": [
                ("Cardiólogo Especialista", 45, 1.3),
                ("Enfermero Especializado", 30, 1.0),
                ("Auxiliar de Enfermería", 20, 1.0)
            ],
            "Implante Marcapasos": [
                ("Cardiólogo Especialista", 120, 1.8),
                ("Enfermero Especializado", 90, 1.3),
                ("Técnico Radiólogo", 60, 1.0),
                ("Auxiliar de Enfermería", 30, 1.0)
            ],
            "Cateterismo Cardíaco Derecho": [
                ("Cardiólogo Especialista", 60, 1.4),
                ("Enfermero Especializado", 50, 1.2),
                ("Técnico Radiólogo", 40, 1.0)
            ],
            "Cateterismo Cardíaco Izquierdo": [
                ("Cardiólogo Especialista", 75, 1.5),
                ("Enfermero Especializado", 60, 1.2),
                ("Técnico Radiólogo", 45, 1.0)
            ],
            "Angioplastia Coronaria": [
                ("Cardiólogo Especialista", 150, 2.0),
                ("Cardiólogo General", 90, 1.5),
                ("Enfermero Especializado", 120, 1.3),
                ("Técnico Radiólogo", 90, 1.0)
            ],
            "Ablación por Radiofrecuencia": [
                ("Cardiólogo Especialista", 180, 2.2),
                ("Enfermero Especializado", 150, 1.5),
                ("Técnico Radiólogo", 120, 1.2)
            ],
            "Cierre de CIA": [
                ("Cardiólogo Especialista", 200, 2.5),
                ("Cardiólogo General", 120, 1.5),
                ("Enfermero Especializado", 180, 1.5),
                ("Técnico Radiólogo", 150, 1.3)
            ],
            "Estudio Hemodinámico Completo": [
                ("Cardiólogo Especialista", 100, 1.6),
                ("Enfermero Especializado", 80, 1.3),
                ("Técnico Radiólogo", 60, 1.0)
            ],
            "Biopsia Endomiocárdica": [
                ("Cardiólogo Especialista", 90, 1.7),
                ("Enfermero Especializado", 70, 1.3),
                ("Técnico Radiólogo", 50, 1.0),
                ("Auxiliar de Enfermería", 30, 1.0)
            ],
        }
        
        row = 6  # Iniciar después de headers (row 5)
        for servicio, ecuaciones in ecuaciones_servicios.items():
            # Buscar código del servicio
            codigo = f"SV{list(ecuaciones_servicios.keys()).index(servicio) + 1:03d}"
            
            for grupo, minutos, factor in ecuaciones:
                ws[f'A{row}'] = codigo
                ws[f'B{row}'] = servicio
                ws[f'C{row}'] = grupo
                ws[f'D{row}'] = minutos
                ws[f'E{row}'] = factor
                ws[f'F{row}'] = f"=D{row}*E{row}"  # Minutos ajustados
                
                # Aplicar estilos
                self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'C{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'D{row}'], "input")
                self.aplicar_estilo_celda(ws[f'E{row}'], "input")
                self.aplicar_estilo_celda(ws[f'F{row}'], "resultado")
                
                # Formato
                ws[f'F{row}'].number_format = '#,##0.0'
                
                row += 1
        
        self.ajustar_columnas(ws, {
            'A': 15, 'B': 38, 'C': 30, 'D': 20, 'E': 20, 'F': 20
        })
    
    def crear_hoja_insumos(self):
        """Crea la hoja INSUMOS con costos de materiales"""
        ws = self.wb.create_sheet("INSUMOS")
        
        # Título
        ws['A1'] = "INSUMOS DIRECTOS POR SERVICIO"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:F1')
        
        # Encabezados
        headers = [
            "Código Servicio", "Nombre Servicio", "Tipo Insumo",
            "Cantidad", "Costo Unitario", "Costo Total Insumo"
        ]
        
        row = 3
        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            ws[f'{col}{row}'] = header
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Definir insumos por servicio (solo servicios que requieren insumos)
        insumos_por_servicio = {
            "Ecocardiograma Transesofágico": [
                ("Gel conductor", 1, 5000),
                ("Sonda transesofágica (uso)", 1, 80000),
                ("Material descartable", 1, 15000)
            ],
            "Estudio Electrofisiológico": [
                ("Catéter diagnóstico", 2, 450000),
                ("Electrodos", 6, 25000),
                ("Material quirúrgico", 1, 120000)
            ],
            "Cardioversión Eléctrica": [
                ("Electrodos adhesivos", 2, 35000),
                ("Medicamentos sedación", 1, 80000),
                ("Material descartable", 1, 20000)
            ],
            "Implante Marcapasos": [
                ("Marcapasos (dispositivo)", 1, 8500000),
                ("Cables de estimulación", 2, 650000),
                ("Material quirúrgico", 1, 180000),
                ("Medicamentos", 1, 120000)
            ],
            "Cateterismo Cardíaco Derecho": [
                ("Catéter Swan-Ganz", 1, 380000),
                ("Introductor", 1, 85000),
                ("Medio de contraste", 50, 1200),
                ("Material descartable", 1, 95000)
            ],
            "Cateterismo Cardíaco Izquierdo": [
                ("Catéter diagnóstico", 2, 420000),
                ("Introductor arterial", 1, 95000),
                ("Medio de contraste", 100, 1200),
                ("Material descartable", 1, 110000)
            ],
            "Angioplastia Coronaria": [
                ("Stent coronario", 1.5, 3500000),
                ("Catéter guía", 1, 580000),
                ("Balón de angioplastia", 2, 450000),
                ("Medio de contraste", 150, 1200),
                ("Medicamentos anticoagulantes", 1, 280000),
                ("Material descartable", 1, 250000)
            ],
            "Ablación por Radiofrecuencia": [
                ("Catéter de ablación", 1, 4500000),
                ("Catéteres diagnósticos", 3, 380000),
                ("Electrodos", 8, 25000),
                ("Material descartable", 1, 320000)
            ],
            "Cierre de CIA": [
                ("Dispositivo oclusor", 1, 12000000),
                ("Catéter delivery", 1, 850000),
                ("Catéter diagnóstico", 2, 380000),
                ("Medio de contraste", 80, 1200),
                ("Material descartable", 1, 280000)
            ],
            "Estudio Hemodinámico Completo": [
                ("Catéteres", 3, 420000),
                ("Introductor", 2, 85000),
                ("Medio de contraste", 120, 1200),
                ("Material descartable", 1, 150000)
            ],
            "Biopsia Endomiocárdica": [
                ("Pinza de biopsia", 1, 680000),
                ("Catéter guía", 1, 350000),
                ("Material histológico", 1, 95000),
                ("Material descartable", 1, 120000)
            ],
        }
        
        row = 4
        for servicio, insumos in insumos_por_servicio.items():
            # Buscar código del servicio
            servicios_list = list(insumos_por_servicio.keys())
            # Necesitamos buscar en la lista original de servicios
            idx = self.servicios.index(servicio) + 1 if servicio in self.servicios else 1
            codigo = f"SV{idx:03d}"
            
            for tipo_insumo, cantidad, costo_unitario in insumos:
                ws[f'A{row}'] = codigo
                ws[f'B{row}'] = servicio
                ws[f'C{row}'] = tipo_insumo
                ws[f'D{row}'] = cantidad
                ws[f'E{row}'] = costo_unitario
                ws[f'F{row}'] = f"=D{row}*E{row}"
                
                # Aplicar estilos
                self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'C{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'D{row}'], "input")
                self.aplicar_estilo_celda(ws[f'E{row}'], "input")
                self.aplicar_estilo_celda(ws[f'F{row}'], "resultado")
                
                # Formato
                ws[f'E{row}'].number_format = '$#,##0'
                ws[f'F{row}'].number_format = '$#,##0'
                
                row += 1
        
        self.ajustar_columnas(ws, {
            'A': 15, 'B': 38, 'C': 35, 'D': 12, 'E': 18, 'F': 20
        })
    
    def crear_hoja_costos_indirectos(self):
        """Crea la hoja COSTOS_INDIRECTOS (Auxiliar Contable Clase 7)"""
        ws = self.wb.create_sheet("COSTOS_INDIRECTOS")
        
        # Título
        ws['A1'] = "AUXILIAR DE COSTOS DE PRODUCCIÓN (CLASE 7)"
        ws['A1'].font = Font(name=self.FUENTE_BASE, size=14, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = "Consolidación de Costos Directos e Indirectos (P.U.C.)"
        ws['A2'].font = Font(name=self.FUENTE_BASE, size=10, italic=True)
        ws.merge_cells('A2:F2')
        
        # Encabezados
        headers = ["Código Cuenta", "Concepto", "Sede", "Valor Mensual", "Criterio Distribución", "Tipo Gasto"]
        
        row = 4
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws[f'{col}{row}'] = header
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].alignment = alignment
            ws[f'{col}{row}'].border = border
            
        row = 5
        
        # 1. INTEGRACIÓN COSTOS DIRECTOS POR SEDE (NOMINA E INSUMOS)
        # Esto asegura que sumen exactamente con el consumo real y la nómina pagada
        
        for sede in self.sedes:
            # 7105 - MATERIA PRIMA (INSUMOS)
            # Valor Real = Suma de (Costo Unitario Insumo * Volumen Mes) para esta Sede
            # Usamos Rangos directos de COSTEO_SERVICIOS para evitar errores #REF de tablas
            # Col C: Sede, Col E: Costo Insumos, Col K: Volumen Mes
            # Usamos rangos limitados (ej 5:1000) en SUMPRODUCT para eficiencia y evitar errores con encabezados
            r_sede = "COSTEO_SERVICIOS!$C$5:$C$500"
            r_costo = "COSTEO_SERVICIOS!$E$5:$E$500"
            r_vol = "COSTEO_SERVICIOS!$K$5:$K$500"
            
            ws[f'A{row}'] = "7105"
            ws[f'B{row}'] = "COSTO MATERIA PRIMA (INSUMOS)"
            ws[f'C{row}'] = sede
            ws[f'D{row}'] = f'=SUMPRODUCT(({r_sede}="{sede}")*({r_costo})*({r_vol}))'
            ws[f'E{row}'] = "Consumo Real (Volumen)"
            ws[f'F{row}'] = "Variable"
            
            self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'C{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'D{row}'], "resultado")
            self.aplicar_estilo_celda(ws[f'E{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'F{row}'], "calculo")
            ws[f'D{row}'].number_format = '$#,##0'
            row += 1
            
            # 7205 - MANO DE OBRA DIRECTA
            # Valor Real = Suma de Nómina asignada a esta sede
            # Usamos NOMINA!G:G (Costo Total) y NOMINA!H:H (Sede)
            ws[f'A{row}'] = "7205"
            ws[f'B{row}'] = "COSTO MANO DE OBRA DIRECTA"
            ws[f'C{row}'] = sede
            ws[f'D{row}'] = f'=SUMIFS(NOMINA!$G:$G,NOMINA!$H:$H,"{sede}")'
            ws[f'E{row}'] = "Nómina Directa"
            ws[f'F{row}'] = "Fijo"
            
            self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'C{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'D{row}'], "resultado")
            self.aplicar_estilo_celda(ws[f'E{row}'], "calculo")
            self.aplicar_estilo_celda(ws[f'F{row}'], "calculo")
            ws[f'D{row}'].number_format = '$#,##0'
            row += 1

        
        # 2. COSTOS INDIRECTOS (73) - DESGLOSE POR SEDE BASADO EN PRESUPUESTO
        if not hasattr(self, 'presupuesto_indirectos'):
             self.inicializar_datos_financieros()
             
        total_salas = sum(self.salas_por_sede.values())
        
        for concepto_raw, valor_total in self.presupuesto_indirectos.items():
            try:
                codigo, nombre = concepto_raw.split(" - ")
            except:
                codigo, nombre = "7399", concepto_raw

            for sede in self.sedes:
                # Calcular valor proporcional a la capacidad (salas) de la sede
                salas = self.salas_por_sede.get(sede, 1) # Usar get por seguridad
                factor = salas / total_salas
                valor_sede = valor_total * factor
                
                ws[f'A{row}'] = codigo
                ws[f'B{row}'] = nombre
                ws[f'C{row}'] = sede
                ws[f'D{row}'] = valor_sede
                ws[f'E{row}'] = "Capacidad Instalada (Minutos)"
                ws[f'F{row}'] = "Fijo"
                
                self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'C{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'D{row}'], "input")
                self.aplicar_estilo_celda(ws[f'E{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'F{row}'], "calculo")
                ws[f'D{row}'].number_format = '$#,##0'
                
                row += 1
                 
        # CREAR TABLA OFICIAL
        self.crear_tabla(ws, "TablaIndirectos", f"A4:F{row-1}")
        
        self.ajustar_columnas(ws, {
            'A': 15, 'B': 45, 'C': 25, 'D': 18, 'E': 25, 'F': 15
        })

    def crear_hoja_produccion(self):
        """Crea la hoja PRODUCCION con volúmenes y facturación"""
        ws = self.wb.create_sheet("PRODUCCION")

        # Título
        ws['A1'] = "PRODUCCIÓN MENSUAL Y FACTURACIÓN"
        ws['A1'].font = Font(name=self.FUENTE_BASE, size=14, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:G1')

        # Encabezados
        headers = [
            "Código", "Servicio", "Sede", "Aseguradora",
            "Cantidad", "Valor Facturado Promedio", "Total Facturado"
        ]

        row = 3
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws[f'{col}{row}'] = header
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Valores de facturación por tipo de servicio (AJUSTADOS PARA RENTABILIDAD)
        valores_facturacion = {
            "Diagnóstico No Invasivo": (120000, 350000),      # Antes 80k-250k
            "Diagnóstico Funcional": (250000, 550000),        # Antes 150k-350k
            "Diagnóstico Invasivo": (1500000, 3500000),       # Antes 450k-1.2M
            "Terapéutico": (5500000, 35000000)                # Antes 3.5M-25M
        }

        # Categorías de servicios (sin cambios)
        categorias_servicios = {
            "Ecocardiograma Transtorácico": "Diagnóstico No Invasivo",
            "Ecocardiograma Transesofágico": "Diagnóstico Invasivo",
            "Holter 24 Horas": "Diagnóstico No Invasivo",
            "Holter 48 Horas": "Diagnóstico No Invasivo",
            "Prueba de Esfuerzo": "Diagnóstico Funcional",
            "Electrocardiograma": "Diagnóstico No Invasivo",
            "MAPA 24 Horas": "Diagnóstico No Invasivo",
            "Ecocardiograma Doppler": "Diagnóstico No Invasivo",
            "Ecocardiograma de Estrés": "Diagnóstico Funcional",
            "Tilt Test": "Diagnóstico Funcional",
            "Estudio Electrofisiológico": "Diagnóstico Invasivo",
            "Cardioversión Eléctrica": "Terapéutico",
            "Implante Marcapasos": "Terapéutico",
            "Cateterismo Cardíaco Derecho": "Diagnóstico Invasivo",
            "Cateterismo Cardíaco Izquierdo": "Diagnóstico Invasivo",
            "Angioplastia Coronaria": "Terapéutico",
            "Ablación por Radiofrecuencia": "Terapéutico",
            "Cierre de CIA": "Terapéutico",
            "Estudio Hemodinámico Completo": "Diagnóstico Invasivo",
            "Biopsia Endomiocárdica": "Diagnóstico Invasivo",
        }

        row = 4
        for idx, servicio in enumerate(self.servicios, 1):
            codigo = f"SV{idx:03d}"
            categoria = categorias_servicios.get(servicio, "Diagnóstico No Invasivo")
            valor_min, valor_max = valores_facturacion.get(categoria, (100000, 500000))

            for sede in self.sedes:
                for aseguradora in self.aseguradoras:
                    # Variar cantidad según tipo (AJUSTE FINO PARA ELIMINAR SOBRE-EJECUCION DE 5.5x)
                    # Objetivo: ~35,000 minutos totales por sede (80% ocupación)
                    if categoria == "Terapéutico":
                        cantidad = random.randint(1, 3)       # Antes 1-4
                    elif categoria == "Diagnóstico Invasivo":
                        cantidad = random.randint(2, 6)       # Antes 4-10
                    elif categoria == "Diagnóstico Funcional":
                        cantidad = random.randint(10, 20)     # Antes 15-30
                    else:
                        cantidad = random.randint(15, 40)     # Antes 30-80. Drástica reducción.

                    # Valor facturado con variación por aseguradora
                    valor = random.randint(int(valor_min), int(valor_max))

                    ws[f'A{row}'] = codigo
                    ws[f'B{row}'] = servicio
                    ws[f'C{row}'] = sede
                    ws[f'D{row}'] = aseguradora
                    ws[f'E{row}'] = cantidad
                    ws[f'F{row}'] = valor
                    ws[f'G{row}'] = f"=E{row}*F{row}"

                    # Aplicar estilos
                    self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
                    self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
                    self.aplicar_estilo_celda(ws[f'C{row}'], "calculo")
                    self.aplicar_estilo_celda(ws[f'D{row}'], "calculo")
                    self.aplicar_estilo_celda(ws[f'E{row}'], "input")
                    self.aplicar_estilo_celda(ws[f'F{row}'], "input")
                    self.aplicar_estilo_celda(ws[f'G{row}'], "resultado")

                    # Formato
                    ws[f'F{row}'].number_format = '$#,##0'
                    ws[f'G{row}'].number_format = '$#,##0'

                    row += 1

        # Crear Tabla Oficial
        self.crear_tabla(ws, "TablaProduccion", f"A3:G{row-1}")

        # Totales
        ws[f'B{row}'] = "TOTAL FACTURACIÓN MENSUAL"
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'E{row}'] = f"=SUM(E4:E{row-1})"
        ws[f'E{row}'].font = Font(bold=True)
        ws[f'G{row}'] = f"=SUM(G4:G{row-1})"
        ws[f'G{row}'].font = Font(bold=True)
        ws[f'G{row}'].number_format = '$#,##0'
        font, fill, alignment, border = self.crear_estilo_header()
        ws[f'B{row}'].fill = fill
        ws[f'E{row}'].fill = fill
        ws[f'G{row}'].fill = fill
        ws.merge_cells(f'B{row}:D{row}')

        self.ajustar_columnas(ws, {
            'A': 12, 'B': 35, 'C': 22, 'D': 20, 'E': 12, 'F': 18, 'G': 18
        })

    def crear_hoja_costeo_servicios(self):
        """Calcula el costo unitario por servicio usando TASAS CIF REALES."""
        ws = self.wb.create_sheet("COSTEO_SERVICIOS")
        
        # Títulos
        ws['A1'] = "HOJA DE COSTEO UNITARIO POR SERVICIO (TDABC)"
        ws['A1'].font = Font(name=self.FUENTE_BASE, size=14, bold=True, color=self.COLOR_HEADER)
        
        ws['A2'] = "Nota: Costo CIF calculado con Tasa Real basada en Capacidad Instalada por Sede."
        ws['A2'].font = Font(name=self.FUENTE_BASE, size=10, italic=True)

        headers = ["Código", "Servicio", "Sede", "Costo MO Directa", "Costo Insumos", "Costo CIF (Indirecto)", "Costo Unitario Total", "Precio Venta Prom.", "Margen Unitario", "Margen %", "Volumen Mes"]

        # Estilo headers
        row = 4
        for i, header in enumerate(headers):
            col = chr(65 + i)
            ws[f'{col}{row}'] = header
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].alignment = alignment
            ws[f'{col}{row}'].border = border

        # Datos
        row = 5
        for servicio in self.servicios:
            for sede in self.sedes:
                ws[f'A{row}'] = "SV00"
                ws[f'B{row}'] = servicio
                ws[f'C{row}'] = sede
                
                # MO Directa
                minutes_ref = f"SUMIFS(ECUACIONES_TIEMPO!$D:$D,ECUACIONES_TIEMPO!$B:$B,B{row})" 
                cost_min_ref = 2200 
                ws[f'D{row}'] = f"={minutes_ref}*{cost_min_ref}"
                
                # Costo Insumos: Sumar si existe. Usamos rangos directos de INSUMOS!
                # Col F: Costo Total, Col B: Nombre Servicio
                ws[f'E{row}'] = f"=SUMIFS(INSUMOS!$F:$F,INSUMOS!$B:$B,B{row})"
                
                # Costo CIF (Indirecto): Minutos * Tasa CIF Sede
                if hasattr(self, 'tasas_cif_por_sede'):
                    cif_rate = self.tasas_cif_por_sede.get(sede, 1400)
                else: 
                     cif_rate = 1400 
                    
                ws[f'F{row}'] = f"={minutes_ref}*{cif_rate:.2f}"

                # Total Unitario
                ws[f'G{row}'] = f"=D{row}+E{row}+F{row}"
                
                # Precio Venta (Promedio de PRODUCCION)
                # Col F: Valor Facturado Prom, Col B: Servicio, Col C: Sede
                ws[f'H{row}'] = f"=IFERROR(AVERAGEIFS(PRODUCCION!$F:$F,PRODUCCION!$B:$B,B{row},PRODUCCION!$C:$C,C{row}),0)"
                
                # Margen
                ws[f'I{row}'] = f"=H{row}-G{row}"
                ws[f'J{row}'] = f"=IF(H{row}>0,I{row}/H{row},0)"
                
                # Volumen (Suma de PRODUCCION)
                # Col E: Cantidad
                ws[f'K{row}'] = f"=SUMIFS(PRODUCCION!$E:$E,PRODUCCION!$B:$B,B{row},PRODUCCION!$C:$C,C{row})"

                # Estilos
                self.aplicar_estilo_celda(ws[f'A{row}'], "normal")
                self.aplicar_estilo_celda(ws[f'B{row}'], "normal")
                self.aplicar_estilo_celda(ws[f'C{row}'], "normal")
                for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                    self.aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
                    if col == 'J': ws[f'{col}{row}'].number_format = '0.0%'
                    elif col == 'K': ws[f'{col}{row}'].number_format = '#,##0'
                    else: ws[f'{col}{row}'].number_format = '$#,##0'
                
                row += 1

        self.crear_tabla(ws, "TablaCosteo", f"A4:K{row-1}")

        self.ajustar_columnas(ws, {
            'A': 10, 'B': 30, 'C': 20, 'D': 15, 'E': 15, 'F': 15, 
            'G': 18, 'H': 18, 'I': 18, 'J': 12, 'K': 15
        })

    def crear_hoja_resumen_ejecutivo(self):
        """Crea la hoja RESUMEN_EJECUTIVO con análisis consolidado y Conciliación"""
        ws = self.wb.create_sheet("RESUMEN_EJECUTIVO")
        
        # Título principal
        ws['A1'] = "CardioCentro Diagnóstico Integral S.A.S."
        ws['A1'].font = Font(name=self.FUENTE_BASE, size=18, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = "RESUMEN EJECUTIVO - ANÁLISIS DE RENTABILIDAD (Enero)"
        ws['A2'].font = Font(name=self.FUENTE_BASE, size=14, bold=True, color=self.COLOR_HEADER)
        ws.merge_cells('A2:F2')
        
        ws['A3'] = f"Período: {datetime.now().strftime('%B %Y')}"
        ws['A3'].font = Font(name=self.FUENTE_BASE, size=11, italic=True)
        ws.merge_cells('A3:F3')
        
        # SECCIÓN 1: Indicadores Globales
        row = 5
        ws[f'A{row}'] = "INDICADORES GLOBALES DE RESULTADOS"
        font, fill, alignment, border = self.crear_estilo_header()
        ws[f'A{row}'].font = font
        ws[f'A{row}'].fill = fill
        ws.merge_cells(f'A{row}:B{row}')
        
        indicadores = [
            ("Ingresos Operacionales (Facturación)", "=SUM(TablaProduccion[Total Facturado])", '$#,##0'),
            ("Total Costos Asignados (TDABC)", "=SUM(TablaCosteo[Costo Unitario Total]*TablaCosteo[Volumen Mes])", '$#,##0'),
            ("   - Costo Personal Directo", "=SUM(TablaCosteo[Costo MO Directa]*TablaCosteo[Volumen Mes])", '$#,##0'),
            ("   - Costo Insumos", "=SUM(TablaCosteo[Costo Insumos]*TablaCosteo[Volumen Mes])", '$#,##0'),
            ("   - Costos Indirectos (CIF)", "=SUM(TablaCosteo[Costo CIF (Indirecto)]*TablaCosteo[Volumen Mes])", '$#,##0'),
            ("Utilidad Operacional (Estimada)", "=B7-B8", '$#,##0'),
            ("Margen Operacional %", "=IF(B7>0,B12/B7,0)", '0.0%'),
            ("Total Servicios Prestados", "=SUM(TablaProduccion[Cantidad])", '#,##0'),
            ("MARGEN OPERATIVO", "=B7-B8", '$#,##0'),
            ("MARGEN OPERATIVO %", "=IF(B7>0,B13/B7,0)", '0.0%'),
        ]
        
        row += 1
        ws[f'A{row}'] = "Indicador"
        ws[f'B{row}'] = "Valor"
        ws[f'C{row}'] = ""
        for col in ['A', 'B']:
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
        
        for indicador, formula, formato in indicadores:
            row += 1
            ws[f'A{row}'] = indicador
            ws[f'B{row}'] = formula
            ws[f'B{row}'].number_format = formato
            
            if "MARGEN" in indicador:
                self.aplicar_estilo_celda(ws[f'A{row}'], "resultado")
                self.aplicar_estilo_celda(ws[f'B{row}'], "resultado")
            else:
                self.aplicar_estilo_celda(ws[f'A{row}'], "calculo")
                self.aplicar_estilo_celda(ws[f'B{row}'], "calculo")
            
            # Sangría visual
            if indicador.startswith("   -"):
                 ws[f'A{row}'].alignment = Alignment(indent=2)


        # SECCIÓN 1.5: CONCILIACIÓN DE COSTOS (CONTABILIDAD UNIFICADA)
        row += 3
        ws[f'A{row}'] = "CONCILIACIÓN DE COSTOS (CONTABLE VS DISTRIBUIDO)"
        font, fill, alignment, border = self.crear_estilo_header()
        ws[f'A{row}'].font = font
        ws[f'A{row}'].fill = fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Concepto (Fuente: COSTOS_INDIRECTOS)"
        ws[f'B{row}'] = "Costo Contable (Gastado)"
        ws[f'C{row}'] = "Costo Asignado (TDABC)"
        
        for col in ['A', 'B', 'C']:
             ws[f'{col}{row}'].font = Font(name=self.FUENTE_BASE, size=10, bold=True)
             ws[f'{col}{row}'].border = border
             
        row += 1
        # Usamos SUMIFS sobre TablaIndirectos para traer los valores contables clasificados
        ws[f'A{row}'] = "Materia Prima / Insumos (71)"
        ws[f'B{row}'] = '=SUMIFS(TablaIndirectos[Valor Mensual],TablaIndirectos[Código Cuenta],"71*")' 
        ws[f'C{row}'] = "=SUMPRODUCT(TablaCosteo[Costo Insumos],TablaCosteo[Volumen Mes])"
        
        row += 1
        ws[f'A{row}'] = "Mano de Obra Directa (72)"
        ws[f'B{row}'] = '=SUMIFS(TablaIndirectos[Valor Mensual],TablaIndirectos[Código Cuenta],"72*")'
        ws[f'C{row}'] = "=SUMPRODUCT(TablaCosteo[Costo MO Directa],TablaCosteo[Volumen Mes])"
        
        row += 1
        ws[f'A{row}'] = "Costos Indirectos CIF (73)"
        ws[f'B{row}'] = '=SUMIFS(TablaIndirectos[Valor Mensual],TablaIndirectos[Código Cuenta],"73*")'
        ws[f'C{row}'] = "=SUMPRODUCT(TablaCosteo[Costo CIF (Indirecto)],TablaCosteo[Volumen Mes])"
        
        # Aplicar formatos
        for r in range(row-2, row+1):
            for col in ['B', 'C']:
                ws[f'{col}{r}'].number_format = '$#,##0'
                self.aplicar_estilo_celda(ws[f'{col}{r}'], "calculo")
            self.aplicar_estilo_celda(ws[f'A{r}'])

        row += 1
        ws[f'A{row}'] = "DIFERENCIA (CAPACIDAD OCIOSA)"
        ws[f'A{row}'].font = Font(name=self.FUENTE_BASE, bold=True, color="FF0000")
        
        # Sumamos las columnas B y C para la diferencia total
        # B actual = row.  Totales = sum(row-3 : row-1)
        r_start = row-3
        r_end = row-1
        ws[f'B{row}'] = f"=SUM(B{r_start}:B{r_end})-SUM(C{r_start}:C{r_end})"
        
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'].alignment = Alignment(horizontal="center")
        ws[f'B{row}'].number_format = '$#,##0'
        
        row += 1
        ws[f'A{row}'] = "NOTA IMPORTANTE: La Diferencia en '73' corresponde a la CAPACIDAD OCIOSA DE INFRAESTRUCTURA.\n   > Valores Positivos (+) indican que se está pagando por una sede que NO se usa completamente (Pérdida por Vacío).\n   > Valores Negativos (-) indican Sobre-ejecución (se trabajó más de la capacidad teórica/turnos extra)."
        ws[f'A{row}'].font = Font(name=self.FUENTE_BASE, italic=True, size=9, color="CC0000")
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 45 # Dar espacio para el texto
        ws.merge_cells(f'A{row}:C{row}')

        
        # SECCIÓN 2: RENTABILIDAD POR SERVICIO (TODOS)
        row += 3
        ws[f'A{row}'] = "RENTABILIDAD POR SERVICIO (TODOS)"
        font, fill, alignment, border = self.crear_estilo_header()
        ws[f'A{row}'].font = font
        ws[f'A{row}'].fill = fill
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws[f'A{row}'] = "Servicio"
        ws[f'B{row}'] = "Costo Total"
        ws[f'C{row}'] = "Facturación Total"
        ws[f'D{row}'] = "Margen Total"
        ws[f'E{row}'] = "Margen %"
        for col in ['A', 'B', 'C', 'D', 'E']:
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
        
        # Listar TODOS los servicios
        for servicio in self.servicios:
            row += 1
            ws[f'A{row}'] = servicio
             # Costo Total = SUMPRODUCT((servicio coincide) * costo total unitario * volumen)
            ws[f'B{row}'] = f"=SUMPRODUCT((TablaCosteo[Servicio]=A{row})*(TablaCosteo[Costo Unitario Total])*(TablaCosteo[Volumen Mes]))"
            # Facturación Total = SUMIF por servicio
            ws[f'C{row}'] = f"=SUMIFS(TablaProduccion[Total Facturado],TablaProduccion[Servicio],A{row})"
            # Margen Total
            ws[f'D{row}'] = f"=C{row}-B{row}"
            # Margen %
            ws[f'E{row}'] = f"=IF(C{row}>0,D{row}/C{row},0)"
            
            for col in ['A']:
                self.aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
            for col in ['B', 'C', 'D', 'E']:
                self.aplicar_estilo_celda(ws[f'{col}{row}'], "resultado")
            
            ws[f'B{row}'].number_format = '$#,##0'
            ws[f'C{row}'].number_format = '$#,##0'
            ws[f'D{row}'].number_format = '$#,##0'
            ws[f'E{row}'].number_format = '0.0%'
        
        # SECCIÓN 3: Análisis por Sede
        row += 3
        ws[f'A{row}'] = "ANÁLISIS POR SEDE"
        font, fill, alignment, border = self.crear_estilo_header()
        ws[f'A{row}'].font = font
        ws[f'A{row}'].fill = fill
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws[f'A{row}'] = "Sede"
        ws[f'B{row}'] = "Facturación"
        ws[f'C{row}'] = "Costos"
        ws[f'D{row}'] = "Margen"
        ws[f'E{row}'] = "Margen %"
        for col in ['A', 'B', 'C', 'D', 'E']:
            font, fill, alignment, border = self.crear_estilo_header()
            ws[f'{col}{row}'].font = font
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = border
        
        for sede in self.sedes:
            row += 1
            ws[f'A{row}'] = sede
            # Facturación por sede
            ws[f'B{row}'] = f'=SUMIFS(TablaProduccion[Total Facturado],TablaProduccion[Sede],A{row})'
            # Costos por sede = SUMPRODUCT de (coincide sede) * costo total unitario * volumen
            ws[f'C{row}'] = f'=SUMPRODUCT((TablaCosteo[Sede]=A{row})*(TablaCosteo[Costo Unitario Total])*(TablaCosteo[Volumen Mes]))'
            # Margen
            ws[f'D{row}'] = f'=B{row}-C{row}'
            # Margen %
            ws[f'E{row}'] = f'=IF(B{row}>0,D{row}/B{row},0)'
            
            for col in ['A']:
                self.aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
            for col in ['B', 'C', 'D', 'E']:
                self.aplicar_estilo_celda(ws[f'{col}{row}'], "resultado")
            
            ws[f'B{row}'].number_format = '$#,##0'
            ws[f'C{row}'].number_format = '$#,##0'
            ws[f'D{row}'].number_format = '$#,##0'
            ws[f'E{row}'].number_format = '0.0%'
        
        # SECCIÓN 4: Notas importantes (Resumidas)
        row += 3
        ws[f'A{row}'] = "NOTAS ADICIONALES:"
        ws[f'A{row}'].font = Font(name=self.FUENTE_BASE, bold=True)
        ws[f'A{row}'].alignment = Alignment(vertical='top')
        ws.merge_cells(f'A{row}:F{row+3}')
        ws[f'A{row}'] = "1. El modelo utiliza referencias directas para cálculo de materiales, asegurando integridad referencial.\n2. La Tabla 'COSTOS_INDIRECTOS' centraliza toda la contabilidad (Insumos 71 + MO 72 + Indirectos 73).\n3. La Capacidad Ociosa refleja la diferencia entre los recursos pagados y los consumidos."
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                         
        self.ajustar_columnas(ws, {
            'A': 45, 'B': 20, 'C': 20, 'D': 20, 'E': 15, 'F': 15
        })
    
    def generar_archivo(self, nombre_archivo="Modelo_TDABC_CardioCentro.xlsx"):
        """Genera el archivo Excel completo"""
        print("Iniciando generación del modelo TDABC...")
        print("="*60)
        
        # Crear todas las hojas
        print("✓ Creando hoja PARAMETROS...")
        self.crear_hoja_parametros()
        
        print("✓ Creando hoja NOMINA...")
        self.crear_hoja_nomina()
        
        print("✓ Creando hoja CAPACIDAD...")
        self.crear_hoja_capacidad()
        
        print("✓ Creando hoja COSTO_POR_MINUTO...")
        self.crear_hoja_costo_por_minuto()
        
        print("✓ Creando hoja SERVICIOS...")
        self.crear_hoja_servicios()
        
        print("✓ Creando hoja ECUACIONES_TIEMPO...")
        self.crear_hoja_ecuaciones_tiempo()
        
        print("✓ Creando hoja INSUMOS...")
        self.crear_hoja_insumos()
        
        print("✓ Creando hoja PRODUCCION...")
        self.crear_hoja_produccion()
        
        print("✓ Creando hoja COSTEO_SERVICIOS...")
        self.crear_hoja_costeo_servicios()
        
        print("✓ Creando hoja COSTOS_INDIRECTOS...")
        self.crear_hoja_costos_indirectos()
        
        print("✓ Creando hoja RESUMEN_EJECUTIVO...")
        self.crear_hoja_resumen_ejecutivo()
        
        print("="*60)
        print(f"Guardando archivo {nombre_archivo}...")
        self.wb.save(nombre_archivo)
        print(f"✓✓✓ Archivo generado exitosamente: {nombre_archivo}")
        print("="*60)
        print("\nESTRUCTURA DEL MODELO:")
        print("  1. PARAMETROS - Configuración general")
        print("  2. NOMINA - Estructura salarial")
        print("  3. CAPACIDAD - Capacidad práctica (184h/mes)")
        print("  4. COSTO_POR_MINUTO - Núcleo TDABC")
        print("  5. SERVICIOS - Catálogo de servicios")
        print("  6. ECUACIONES_TIEMPO - Ecuaciones TDABC por servicio")
        print("  7. INSUMOS - Costos de materiales")
        print("  8. COSTOS_INDIRECTOS - Costos administrativos")
        print("  9. PRODUCCION - Volúmenes y facturación")
        print(" 10. COSTEO_SERVICIOS - Costo total por servicio")
        print(" 11. RESUMEN_EJECUTIVO - Dashboard de rentabilidad")
        print("="*60)



# EJECUTAR GENERACIÓN
if __name__ == "__main__":
    print("\n" + "="*60)
    print("GENERADOR DE MODELO TDABC")
    print("CardioCentro Diagnóstico Integral S.A.S.")
    print("="*60 + "\n")
    
    modelo = ModeloTDABC()
    modelo.generar_archivo("Modelo_TDABC_CardioCentro.xlsx")
    
    print("\n✅ MODELO TDABC GENERADO CORRECTAMENTE")
    print("\n" + "="*60)
