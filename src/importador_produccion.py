"""
Importador de datos reales de producción y facturación.

Este módulo permite importar datos reales desde archivos Excel/CSV
en lugar de generar datos aleatorios.
"""
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any


class ImportadorProduccion:
    """
    Importa datos reales de producción y facturación desde archivos externos.
    """
    
    # Mapeo de columnas del sistema real a columnas estándar
    MAPEO_COLUMNAS_REAL = {
        'codigo_servicio': 'codigo_servicio',
        'servicio': 'nombre_servicio',
        'Sede': 'sede',
        'entidad': 'cliente',
        'Fecha Cita': 'fecha',
        'Valor Servicio': 'valor_unitario',
        'Valor Recaudo': 'valor_recaudo',
        'No Factura': 'numero_factura',
        'Identificacion': 'identificacion_paciente',
        'estado': 'estado',
        'area': 'area'
    }
    
    # Columnas mínimas requeridas (simplificado)
    COLUMNAS_REQUERIDAS = {
        'codigo_servicio': 'Código del servicio prestado',
        'servicio': 'Nombre del servicio (o nombre_servicio)',
        'Sede': 'Sede donde se prestó el servicio',
        'entidad': 'EPS/Aseguradora/Cliente',
        'Valor Servicio': 'Valor facturado del servicio'
    }
    
    def __init__(self, ruta_archivo: str = None):
        """
        Inicializa el importador.
        
        Args:
            ruta_archivo: Ruta al archivo Excel/CSV con datos de producción
        """
        self.ruta_archivo = ruta_archivo
        self.datos_produccion = None
        self.errores = []
        self.advertencias = []
    
    def cargar_desde_excel(self, ruta_archivo: str, hoja: str = None) -> bool:
        """
        Carga datos de producción desde un archivo Excel.
        
        Args:
            ruta_archivo: Ruta al archivo Excel
            hoja: Nombre de la hoja (opcional, usa la primera si no se especifica)
            
        Returns:
            True si la carga fue exitosa, False en caso contrario
        """
        try:
            if hoja:
                df = pd.read_excel(ruta_archivo, sheet_name=hoja)
            else:
                df = pd.read_excel(ruta_archivo)
            
            return self._procesar_dataframe(df, ruta_archivo)
        
        except FileNotFoundError:
            self.errores.append(f"Archivo no encontrado: {ruta_archivo}")
            return False
        except Exception as e:
            self.errores.append(f"Error al leer Excel: {e}")
            return False
    
    def cargar_desde_csv(self, ruta_archivo: str, separador: str = ',') -> bool:
        """
        Carga datos de producción desde un archivo CSV.
        
        Args:
            ruta_archivo: Ruta al archivo CSV
            separador: Separador de columnas (por defecto ',')
            
        Returns:
            True si la carga fue exitosa, False en caso contrario
        """
        try:
            df = pd.read_csv(ruta_archivo, sep=separador, encoding='utf-8')
            return self._procesar_dataframe(df, ruta_archivo)
        
        except FileNotFoundError:
            self.errores.append(f"Archivo no encontrado: {ruta_archivo}")
            return False
        except Exception as e:
            self.errores.append(f"Error al leer CSV: {e}")
            return False
    
    def _procesar_dataframe(self, df: pd.DataFrame, origen: str) -> bool:
        """
        Procesa y valida el DataFrame cargado.
        
        Args:
            df: DataFrame con datos de producción
            origen: Nombre del archivo de origen
            
        Returns:
            True si el procesamiento fue exitoso
        """
        # Validar columnas requeridas
        columnas_faltantes = []
        for col_requerida in self.COLUMNAS_REQUERIDAS.keys():
            if col_requerida not in df.columns:
                columnas_faltantes.append(col_requerida)
        
        if columnas_faltantes:
            self.errores.append(
                f"Columnas faltantes en {origen}: {', '.join(columnas_faltantes)}"
            )
            self.errores.append(
                f"Columnas requeridas: {', '.join(self.COLUMNAS_REQUERIDAS.keys())}"
            )
            return False
        
        # Validar datos
        if len(df) == 0:
            self.errores.append(f"El archivo {origen} no contiene datos")
            return False
        
        # Validar tipos de datos
        errores_validacion = []
        
        for idx, row in df.iterrows():
            # Validar cantidad
            if pd.isna(row['cantidad']) or row['cantidad'] <= 0:
                errores_validacion.append(
                    f"Fila {idx+2}: cantidad debe ser mayor a 0"
                )
            
            # Validar valor_unitario
            if pd.isna(row['valor_unitario']) or row['valor_unitario'] <= 0:
                errores_validacion.append(
                    f"Fila {idx+2}: valor_unitario debe ser mayor a 0"
                )
            
            # Validar campos obligatorios
            if pd.isna(row['codigo_servicio']) or str(row['codigo_servicio']).strip() == '':
                errores_validacion.append(
                    f"Fila {idx+2}: codigo_servicio es obligatorio"
                )
            
            if pd.isna(row['sede']) or str(row['sede']).strip() == '':
                errores_validacion.append(
                    f"Fila {idx+2}: sede es obligatoria"
                )
            
            if pd.isna(row['cliente']) or str(row['cliente']).strip() == '':
                errores_validacion.append(
                    f"Fila {idx+2}: cliente es obligatorio"
                )
        
        if errores_validacion:
            self.errores.extend(errores_validacion[:10])  # Mostrar solo primeros 10
            if len(errores_validacion) > 10:
                self.errores.append(
                    f"... y {len(errores_validacion) - 10} errores más"
                )
            return False
        
        # Guardar datos procesados
        self.datos_produccion = df.to_dict('records')
        
        # Generar advertencias informativas
        total_registros = len(df)
        sedes_unicas = df['sede'].nunique()
        clientes_unicos = df['cliente'].nunique()
        servicios_unicos = df['codigo_servicio'].nunique()
        
        self.advertencias.append(f"Datos cargados exitosamente:")
        self.advertencias.append(f"  - Total registros: {total_registros}")
        self.advertencias.append(f"  - Sedes: {sedes_unicas}")
        self.advertencias.append(f"  - Clientes/EPS: {clientes_unicos}")
        self.advertencias.append(f"  - Servicios diferentes: {servicios_unicos}")
        
        return True
    
    def obtener_datos_produccion(self) -> List[Dict[str, Any]]:
        """
        Obtiene los datos de producción procesados.
        
        Returns:
            Lista de diccionarios con datos de producción
        """
        if not self.datos_produccion:
            return []
        
        return [
            {
                'codigo': str(row['codigo_servicio']),
                'servicio': row.get('nombre_servicio', row['codigo_servicio']),
                'sede': str(row['sede']),
                'aseguradora': str(row['cliente']),
                'cantidad': int(row['cantidad']),
                'valor_unitario': float(row['valor_unitario']),
                'categoria': row.get('categoria', '')
            }
            for row in self.datos_produccion
        ]
    
    def generar_plantilla_excel(self, ruta_salida: str):
        """
        Genera un archivo Excel plantilla con las columnas requeridas.
        
        Args:
            ruta_salida: Ruta donde guardar la plantilla
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos Producción"
        
        # Encabezados
        headers = list(self.COLUMNAS_REQUERIDAS.keys())
        descripciones = list(self.COLUMNAS_REQUERIDAS.values())
        
        # Fila 1: Descripciones
        for col_idx, desc in enumerate(descripciones, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = desc
            cell.font = Font(italic=True, size=9, color="666666")
            cell.alignment = Alignment(wrap_text=True)
        
        # Fila 2: Nombres de columnas
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Fila 3: Ejemplo
        ejemplo = [
            "SV001",
            "Ecocardiograma Transtorácico",
            "Sede Principal",
            "EPS Sura",
            25,
            150000
        ]
        
        for col_idx, valor in enumerate(ejemplo, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = valor
            cell.font = Font(italic=True, color="999999")
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        
        # Guardar
        wb.save(ruta_salida)
        print(f"[OK] Plantilla generada: {ruta_salida}")
        print(f"  1. Llena la plantilla con tus datos reales")
        print(f"  2. Guarda el archivo")
        print(f"  3. Importa con: importador.cargar_desde_excel('{ruta_salida}')")
    
    def obtener_reporte(self) -> str:
        """
        Genera un reporte de la importación.
        
        Returns:
            String con el reporte formateado
        """
        reporte = []
        reporte.append("=" * 60)
        reporte.append("REPORTE DE IMPORTACIÓN DE DATOS DE PRODUCCIÓN")
        reporte.append("=" * 60)
        
        if self.errores:
            reporte.append(f"\n[ERROR] ERRORES ({len(self.errores)}):")
            for error in self.errores:
                reporte.append(f"  - {error}")
        
        if self.advertencias:
            reporte.append(f"\n[INFO] INFORMACIÓN:")
            for adv in self.advertencias:
                reporte.append(f"  {adv}")
        
        if not self.errores and not self.advertencias:
            reporte.append("\n[WARN] No se han cargado datos aún")
        
        reporte.append("=" * 60)
        return "\n".join(reporte)


def ejemplo_uso():
    """Ejemplo de cómo usar el importador"""
    print("\n" + "="*60)
    print("EJEMPLO: Importar Datos Reales de Producción")
    print("="*60 + "\n")
    
    # Paso 1: Generar plantilla
    importador = ImportadorProduccion()
    importador.generar_plantilla_excel("plantilla_produccion.xlsx")
    
    print("\n" + "="*60)
    print("INSTRUCCIONES:")
    print("="*60)
    print("\n1. Abre 'plantilla_produccion.xlsx'")
    print("2. Llena con tus datos REALES de facturación:")
    print("   - Código del servicio")
    print("   - Nombre del servicio")
    print("   - Sede donde se prestó")
    print("   - EPS/Cliente que pagó")
    print("   - Cantidad EXACTA prestada")
    print("   - Valor EXACTO facturado")
    print("\n3. Guarda el archivo")
    print("\n4. Importa los datos:")
    print("   >>> from src.importador_produccion import ImportadorProduccion")
    print("   >>> importador = ImportadorProduccion()")
    print("   >>> importador.cargar_desde_excel('plantilla_produccion.xlsx')")
    print("   >>> datos = importador.obtener_datos_produccion()")
    print("\n" + "="*60 + "\n")


if __name__ == "__main__":
    ejemplo_uso()
