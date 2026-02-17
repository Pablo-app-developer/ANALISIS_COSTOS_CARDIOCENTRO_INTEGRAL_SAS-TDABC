"""
Generador de la hoja CAPACIDAD
"""
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas


def crear_hoja_capacidad(wb):
    """Crea la hoja CAPACIDAD con cálculo de capacidad práctica"""
    ws = wb.create_sheet("CAPACIDAD")
    
    # Título
    ws['A1'] = "ANÁLISIS DE CAPACIDAD PRÁCTICA"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A1:G1')
    
    # Explicación de capacidad práctica
    ws['A2'] = "NOTA: La capacidad práctica (85%) es el tiempo REAL disponible. Se descuenta ~15% por pausas, reuniones, capacitación, etc."
    ws['A2'].font = Font(name='Calibri', size=9, italic=True, color="D35400")
    ws['A2'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A2:G2')
    
    # Encabezados
    headers = [
        "Grupo Ocupacional", "Horas Mensuales", "Minutos Disponibles",
        "% Capacidad Práctica", "Minutos Capacidad Práctica",
        "Personal Disponible", "Total Minutos Grupo"
    ]
    
    row = 4
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Datos de capacidad
    row = 5  # Iniciar después de los headers (row 4)
    for grupo, _ in config.GRUPOS_OCUPACIONALES:
        ws[f'A{row}'] = grupo
        ws[f'B{row}'] = "=PARAMETROS!$B$7"  # Horas mensuales (184h)
        ws[f'C{row}'] = f"=B{row}*60"  # Minutos disponibles
        ws[f'D{row}'] = "=PARAMETROS!$B$9/100"  # % Capacidad práctica (85%)
        ws[f'E{row}'] = f"=C{row}*D{row}"  # Minutos capacidad práctica
        # Sumar personal de este grupo en todas las sedes
        ws[f'F{row}'] = f'=SUMIF(NOMINA!$B:$B,A{row},NOMINA!$F:$F)'
        ws[f'G{row}'] = f"=E{row}*F{row}"  # Total minutos grupo
        
        # Aplicar estilos
        aplicar_estilo_celda(ws[f'A{row}'], "calculo")
        aplicar_estilo_celda(ws[f'B{row}'], "calculo")
        aplicar_estilo_celda(ws[f'C{row}'], "calculo")
        aplicar_estilo_celda(ws[f'D{row}'], "calculo")
        aplicar_estilo_celda(ws[f'E{row}'], "resultado")
        aplicar_estilo_celda(ws[f'F{row}'], "calculo")
        aplicar_estilo_celda(ws[f'G{row}'], "resultado")
        
        # Formato
        ws[f'D{row}'].number_format = '0%'
        ws[f'E{row}'].number_format = '#,##0'
        ws[f'G{row}'].number_format = '#,##0'
        
        row += 1
    
    # Totales
    ws[f'A{row}'] = "TOTAL CAPACIDAD ORGANIZACIONAL"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=SUM(G4:G{row-1})"
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'G{row}'].number_format = '#,##0'
    font, fill, alignment, border = crear_estilo_header()
    ws[f'A{row}'].fill = fill
    ws[f'G{row}'].fill = fill
    ws.merge_cells(f'A{row}:F{row}')
    
    ajustar_columnas(ws, {
        'A': 30, 'B': 18, 'C': 20, 'D': 20,
        'E': 25, 'F': 20, 'G': 25
    })
