"""
Generador de la hoja COSTEO_SERVICIOS
"""
from openpyxl.styles import Font
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas
from ..utils import crear_tabla


def crear_hoja_costeo_servicios(wb, data_init):
    """Calcula el costo unitario por servicio usando TASAS CIF REALES."""
    ws = wb.create_sheet("COSTEO_SERVICIOS")
    
    # Títulos
    ws['A1'] = "HOJA DE COSTEO UNITARIO POR SERVICIO (TDABC)"
    ws['A1'].font = Font(name=config.FUENTE_BASE, size=14, bold=True, color=config.COLOR_HEADER)
    
    ws['A2'] = "Nota: Costo CIF calculado con Tasa Real basada en Capacidad Instalada por Sede."
    ws['A2'].font = Font(name=config.FUENTE_BASE, size=10, italic=True)

    # Encabezados dinámicos desde el mapeo
    _m = config._mapper.get_columnas_estandar("costeo_servicios")
    headers = [
        _m.get("codigo", "Código"),
        _m.get("servicio", "Servicio"),
        _m.get("centro", "Sede"),
        _m.get("mo", "Costo MO Directa"),
        _m.get("insumos", "Costo Insumos"),
        _m.get("cif", "Costo CIF (Indirecto)"),
        _m.get("total", "Costo Unitario Total"),
        _m.get("precio", "Precio Venta Prom."),
        _m.get("margen_moneda", "Margen Unitario"),
        _m.get("margen_pct", "Margen %"),
        _m.get("volumen", "Volumen Mes")
    ]

    # Estilo headers
    row = 4
    for i, header in enumerate(headers):
        col = chr(65 + i)
        ws[f'{col}{row}'] = header
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].alignment = alignment
        ws[f'{col}{row}'].border = border

    # Datos
    row = 5
    for servicio in config.SERVICIOS:
        for sede in config.SEDES:
            ws[f'A{row}'] = "SV00"
            ws[f'B{row}'] = servicio
            ws[f'C{row}'] = sede
            
            # Definir referencia de minutos (necesaria para CIF)
            minutes_ref = f"SUMIFS(ECUACIONES_TIEMPO!$D:$D,ECUACIONES_TIEMPO!$B:$B,B{row})"
            
            # MO Directa: Suma del Costo MO Total calculado en ECUACIONES_TIEMPO (Columna H)
            # Esto considera el mix exacto de especialistas vs técnicos para cada servicio
            ws[f'D{row}'] = f"=SUMIFS(ECUACIONES_TIEMPO!$H:$H,ECUACIONES_TIEMPO!$B:$B,B{row})"
            
            # Costo Insumos: Sumar si existe
            ws[f'E{row}'] = f"=SUMIFS(INSUMOS!$F:$F,INSUMOS!$B:$B,B{row})"
            
            # Costo CIF (Indirecto): Minutos * Tasa CIF Real calculada por Sede
            # Buscamos la tasa en la tabla de tasas al final de COSTOS_INDIRECTOS
            # Usamos SUMIFS en Col A (Sede) y Col D (Tasa), ya que en la tabla principal Col A son Cuentas
            cif_rate_ref = f"SUMIFS(COSTOS_INDIRECTOS!$D:$D,COSTOS_INDIRECTOS!$A:$A,C{row})"
            ws[f'F{row}'] = f"={minutes_ref}*{cif_rate_ref}"

            # Total Unitario
            ws[f'G{row}'] = f"=D{row}+E{row}+F{row}"
            
            # Precio Venta (Promedio de PRODUCCION)
            ws[f'H{row}'] = f"=IFERROR(AVERAGEIFS(PRODUCCION!$F:$F,PRODUCCION!$B:$B,B{row},PRODUCCION!$C:$C,C{row}),0)"
            
            # Margen
            ws[f'I{row}'] = f"=H{row}-G{row}"
            ws[f'J{row}'] = f"=IF(H{row}>0,I{row}/H{row},0)"
            
            # Volumen (Suma de PRODUCCION)
            ws[f'K{row}'] = f"=SUMIFS(PRODUCCION!$E:$E,PRODUCCION!$B:$B,B{row},PRODUCCION!$C:$C,C{row})"

            # Estilos
            aplicar_estilo_celda(ws[f'A{row}'], "normal")
            aplicar_estilo_celda(ws[f'B{row}'], "normal")
            aplicar_estilo_celda(ws[f'C{row}'], "normal")
            for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
                if col == 'J':
                    ws[f'{col}{row}'].number_format = '0.0%'
                elif col == 'K':
                    ws[f'{col}{row}'].number_format = '#,##0'
                else:
                    ws[f'{col}{row}'].number_format = '$#,##0'
            
            row += 1

    crear_tabla(ws, "TablaCosteo", f"A4:K{row-1}")

    ajustar_columnas(ws, {
        'A': 10, 'B': 30, 'C': 20, 'D': 15, 'E': 15, 'F': 15, 
        'G': 18, 'H': 18, 'I': 18, 'J': 12, 'K': 15
    })
