"""
Generador de la hoja COSTO_POR_MINUTO
"""
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas
from ..utils import crear_tabla


def crear_hoja_costo_por_minuto(wb):
    """Crea la hoja COSTO_POR_MINUTO"""
    ws = wb.create_sheet("COSTO_POR_MINUTO")
    
    # Título
    ws['A1'] = "COSTO POR MINUTO - NÚCLEO DEL MODELO TDABC"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A1:E1')
    
    # Encabezados
    headers = [
        "Grupo Ocupacional", "Costo Total Mensual", 
        "Minutos Capacidad Práctica", "Costo por Minuto", "Sede"
    ]
    
    row = 3
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Datos
    row = 4
    for grupo, _ in config.GRUPOS_OCUPACIONALES:
        for sede in config.SEDES:
            ws[f'A{row}'] = grupo
            # Buscar costo total mensual de este grupo en esta sede
            ws[f'B{row}'] = f'=SUMIFS(NOMINA!$E:$E,NOMINA!$B:$B,A{row},NOMINA!$H:$H,E{row})'
            # Minutos capacidad práctica por persona
            ws[f'C{row}'] = f'=VLOOKUP(A{row},CAPACIDAD!$A:$E,5,FALSE)'
            # Costo por minuto = Costo Total / Minutos Capacidad Práctica
            ws[f'D{row}'] = f'=IF(C{row}>0,B{row}/C{row},0)'
            ws[f'E{row}'] = sede
            
            # Aplicar estilos
            aplicar_estilo_celda(ws[f'A{row}'], "calculo")
            aplicar_estilo_celda(ws[f'B{row}'], "calculo")
            aplicar_estilo_celda(ws[f'C{row}'], "calculo")
            aplicar_estilo_celda(ws[f'D{row}'], "resultado")
            aplicar_estilo_celda(ws[f'E{row}'], "calculo")
            
            # Formato
            ws[f'B{row}'].number_format = '$#,##0'
            ws[f'C{row}'].number_format = '#,##0'
            ws[f'D{row}'].number_format = '$#,##0.00'
            
            row += 1
    
    row += 1
    
    # Crear Tabla Estructurada para facilitar Tablas Dinámicas manuales
    # Rango A3:E{row-1} (Desde headers hasta última fila de datos)
    crear_tabla(ws, "TablaDetalleCostoMinuto", f"A3:E{row-1}")
    
    row += 2
    
    # ==============================================================================
    # TABLA MAESTRA CRUZADA (Simulación de Tabla Dinámica)
    # Filas: Grupos Ocupacionales
    # Columnas: Promedio Nacional (H) + Detalle por Sede (I, J, K...)
    # ==============================================================================
    
    ws['G3'] = "TABLA DINÁMICA DE COSTOS (Costo Minuto Real)"
    ws['G3'].font = Font(bold=True, size=11, color="2E86C1")
    
    # Encabezados
    headers_crosstab = ["Grupo Ocupacional", "Promedio Nacional"] + config.SEDES
    
    for i, h in enumerate(headers_crosstab):
        col = get_column_letter(7 + i) # G=7
        celda = ws[f'{col}4']
        celda.value = h
        estilo = crear_estilo_header()
        celda.font = estilo[0]
        celda.fill = estilo[1]
        celda.alignment = Alignment(horizontal="center", wrap_text=True)
        
    row = 5
    for grupo, _ in config.GRUPOS_OCUPACIONALES:
        ws[f'G{row}'] = grupo
        aplicar_estilo_celda(ws[f'G{row}'], "normal")
        
        # Col H: Promedio Nacional (Mantenemos compatibilidad con ECUACIONES_TIEMPO)
        ws[f'H{row}'] = f"=IFERROR(AVERAGEIFS(TablaDetalleCostoMinuto[Costo por Minuto], TablaDetalleCostoMinuto[Grupo Ocupacional], G{row}), 0)"
        ws[f'H{row}'].number_format = '$#,##0.00'
        aplicar_estilo_celda(ws[f'H{row}'], "resultado")
        
        # Cols I...: Costo por Sede Específica
        for i, sede in enumerate(config.SEDES):
            col = get_column_letter(9 + i) # I=9
            # Buscamos el costo específico interceptando Grupo y Sede
            # SUMIFS(Costo, Grupo=G, Sede=Header)
            # Como es único por sede/grupo, SUMIFS o AVERAGEIFS da lo mismo
            ws[f'{col}{row}'] = f"=SUMIFS(TablaDetalleCostoMinuto[Costo por Minuto], TablaDetalleCostoMinuto[Grupo Ocupacional], G{row}, TablaDetalleCostoMinuto[Sede], \"{sede}\")"
            ws[f'{col}{row}'].number_format = '$#,##0.00'
            aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
            
        row += 1

    # Ajustar columnas de la tabla cruzada
    anchos = {'G': 35, 'H': 20}
    for i in range(len(config.SEDES)):
        col = get_column_letter(9 + i)
        anchos[col] = 18
    ajustar_columnas(ws, anchos)
