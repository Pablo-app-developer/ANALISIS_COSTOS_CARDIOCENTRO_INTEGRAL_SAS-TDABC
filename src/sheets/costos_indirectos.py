"""
Generador de la hoja COSTOS_INDIRECTOS
"""
from openpyxl.styles import Font
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas
from ..utils import crear_tabla


def crear_hoja_costos_indirectos(wb, data_init):
    """Crea la hoja COSTOS_INDIRECTOS (Auxiliar Contable Clase 7)"""
    ws = wb.create_sheet("COSTOS_INDIRECTOS")
    
    # Título
    ws['A1'] = "AUXILIAR DE COSTOS DE PRODUCCIÓN (CLASE 7)"
    ws['A1'].font = Font(name=config.FUENTE_BASE, size=14, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Consolidación de Costos Directos e Indirectos (P.U.C.)"
    ws['A2'].font = Font(name=config.FUENTE_BASE, size=10, italic=True)
    ws.merge_cells('A2:F2')
    
    # Encabezados dinámicos desde el mapeo
    _m = config._mapper.get_columnas_estandar("contabilidad")
    headers = [
        _m.get("cuenta", "Código Cuenta"),
        _m.get("nombre", "Concepto"),
        _m.get("centro", "Sede"),
        _m.get("valor", "Valor Mensual"),
        _m.get("criterio", "Criterio Distribución"),
        _m.get("tipo", "Tipo Gasto")
    ]
    
    row = 4
    for i, header in enumerate(headers):
        col = chr(65 + i)
        ws[f'{col}{row}'] = header
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].alignment = alignment
        ws[f'{col}{row}'].border = border
        
    row = 5
    
    # Obtener cuentas directas desde el plan contable
    cuenta_mp = config._mapper.get_cuenta_materia_prima()
    cuenta_mo = config._mapper.get_cuenta_mano_obra()
    
    # 1. INTEGRACIÓN COSTOS DIRECTOS POR SEDE (NOMINA E INSUMOS)
    for sede in config.SEDES:
        # MATERIA PRIMA (INSUMOS)
        r_sede = "COSTEO_SERVICIOS!$C$5:$C$500"
        r_costo = "COSTEO_SERVICIOS!$E$5:$E$500"
        r_vol = "COSTEO_SERVICIOS!$K$5:$K$500"
        
        ws[f'A{row}'] = cuenta_mp.get("codigo", "7105")
        ws[f'B{row}'] = cuenta_mp.get("nombre", "COSTO MATERIA PRIMA (INSUMOS)")
        ws[f'C{row}'] = sede
        ws[f'D{row}'] = f'=SUMPRODUCT(({r_sede}="{sede}")*({r_costo})*({r_vol}))'
        ws[f'E{row}'] = "Consumo Real (Volumen)"
        ws[f'F{row}'] = cuenta_mp.get("tipo", "Variable")
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            if col == 'D':
                 aplicar_estilo_celda(ws[f'{col}{row}'], "resultado")
            else:
                 aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
        ws[f'D{row}'].number_format = '$#,##0'
        row += 1
        
        # MANO DE OBRA DIRECTA
        ws[f'A{row}'] = cuenta_mo.get("codigo", "7205")
        ws[f'B{row}'] = cuenta_mo.get("nombre", "COSTO MANO DE OBRA DIRECTA")
        ws[f'C{row}'] = sede
        ws[f'D{row}'] = f'=SUMIFS(NOMINA!$G:$G,NOMINA!$H:$H,"{sede}")'
        ws[f'E{row}'] = "Nómina Directa"
        ws[f'F{row}'] = cuenta_mo.get("tipo", "Fijo")
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            if col == 'D':
                 aplicar_estilo_celda(ws[f'{col}{row}'], "resultado")
            else:
                 aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
        ws[f'D{row}'].number_format = '$#,##0'
        row += 1

    # 2. COSTOS INDIRECTOS (73) - DESGLOSE POR SEDE BASADO EN PRESUPUESTO
    total_salas = sum(data_init.salas_por_sede.values())
    
    # Obtener información extendida de costos indirectos si existe
    indirectos_config = config._mapper.get_costos_indirectos()
    mapping_indirectos = {f"{c['codigo']} - {c['nombre']}": c for c in indirectos_config} if indirectos_config else {}

    for concepto_raw, valor_total in data_init.presupuesto_indirectos.items():
        try:
            codigo, nombre = concepto_raw.split(" - ")
        except:
            codigo, nombre = "7399", concepto_raw
        
        extra_info = mapping_indirectos.get(concepto_raw, {})
        criterio = extra_info.get("criterio_distribucion", "Capacidad Instalada (Minutos)")
        tipo_gasto = extra_info.get("tipo", "Fijo")

        for sede in config.SEDES:
            # Calcular valor proporcional a la capacidad (salas) de la sede
            salas = data_init.salas_por_sede.get(sede, 1)
            factor = salas / total_salas
            valor_sede = valor_total * factor
            
            ws[f'A{row}'] = codigo
            ws[f'B{row}'] = nombre
            ws[f'C{row}'] = sede
            ws[f'D{row}'] = valor_sede
            ws[f'E{row}'] = criterio
            ws[f'F{row}'] = tipo_gasto
            
            for col in ['A', 'B', 'C', 'E', 'F']:
                 aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
            aplicar_estilo_celda(ws[f'D{row}'], "input")
            ws[f'D{row}'].number_format = '$#,##0'
            
            row += 1
             
    # CREAR TABLA OFICIAL
    crear_tabla(ws, "TablaIndirectos", f"A4:F{row-1}")
    
    row += 2
    
    # 3. CÁLCULO DE TASAS CIF REALES POR SEDE
    ws[f'A{row}'] = "CÁLCULO DE TASAS CIF REALES POR SEDE"
    ws[f'A{row}'].font = Font(bold=True, size=11, color="2E86C1")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers_cif = ["Sede", "Total Costos Indirectos (73)", "Capacidad Minutos Mes", "Tasa CIF Real / Min"]
    for i, h in enumerate(headers_cif):
        cell = ws.cell(row=row, column=i+1, value=h)
        estilo_header = crear_estilo_header()
        cell.font = estilo_header[0]
        cell.fill = estilo_header[1]
    
    row_inicio_tasas = row + 1
    
    for sede in data_init.salas_por_sede.keys():
        row += 1
        ws[f'A{row}'] = sede
        
        # Sumar solo cuentas que empiezan con "73" para esta sede
        # Formula: SUMIFS(TablaIndirectos[Valor], TablaIndirectos[Sede], Sede, TablaIndirectos[Cuenta], "73*")
        # Nota: Como "73*" es texto, funciona bien.
        ws[f'B{row}'] = f'=SUMIFS(D:D, C:C, A{row}, A:A, "73*")'
        
        # Capacidad: Buscar en hoja CAPACIDAD o estimar.
        # CAPACIDAD tiene una tabla, pero no un resumen simple por sede.
        # Asumiremos la capacidad estándar de la empresa * número de salas
        salas = data_init.salas_por_sede.get(sede, 1)
        # Capacidad estándar = 11040 min/mes * salas?
        # No, la capacidad es por RECURSO HUMANO generalmente, pero para CIF se suele usar
        # capacidad instalada de planta o suma de tiempos de personal.
        # Usaremos Capacidad Práctica Mensual Estándar * Salas como driver simple.
        # Mejor: Sumar capacidad de todo el personal de esa sede en CAPACIDAD?
        # Para simplificar y ser robusto: Usar un Named Range o valor fijo calculado aqui.
        capacidad_base = config.CAPACIDAD_MENSUAL_MINUTOS
        ws[f'C{row}'] = f"={capacidad_base}*{salas}" # Minutos disponibles por sala
        
        # Tasa CIF
        ws[f'D{row}'] = f'=IF(C{row}>0,B{row}/C{row},0)'
        
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'C{row}'].number_format = '#,##0'
        ws[f'D{row}'].number_format = '$#,##0.00'
        
        aplicar_estilo_celda(ws[f'A{row}'], "normal")
        aplicar_estilo_celda(ws[f'B{row}'], "calculo")
        aplicar_estilo_celda(ws[f'C{row}'], "calculo")
        aplicar_estilo_celda(ws[f'D{row}'], "resultado")

    ajustar_columnas(ws, {
        'A': 15, 'B': 45, 'C': 25, 'D': 18, 'E': 25, 'F': 15
    })
