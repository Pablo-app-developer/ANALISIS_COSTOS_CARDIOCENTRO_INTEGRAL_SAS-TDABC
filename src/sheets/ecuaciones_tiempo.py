"""
Generador de la hoja ECUACIONES_TIEMPO
"""
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas
from ..data.ecuaciones_data import ECUACIONES_SERVICIOS


def crear_hoja_ecuaciones_tiempo(wb):
    """Crea la hoja ECUACIONES_TIEMPO - núcleo del TDABC"""
    ws = wb.create_sheet("ECUACIONES_TIEMPO")
    
    # Título
    ws['A1'] = "ECUACIONES DE TIEMPO TDABC"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Núcleo del modelo: Cada servicio consume tiempo de diferentes grupos ocupacionales"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    ws.merge_cells('A2:F2')
    
    ws['A3'] = "FACTOR DE COMPLEJIDAD: Ajusta el tiempo según dificultad del caso. 1.0=normal, 1.2-1.5=complejo, 2.0+=muy complejo. Es AJUSTABLE según su realidad."
    ws['A3'].font = Font(name='Calibri', size=9, italic=True, color="D35400")
    ws['A3'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A3:F3')
    
    # Encabezados dinámicos desde el mapeo
    _m = config._mapper.get_columnas_estandar("ecuaciones_tiempo")
    headers = [
        _m.get("codigo_servicio", "Código Servicio"),
        _m.get("nombre_servicio", "Nombre Servicio"),
        _m.get("grupo", "Grupo Ocupacional"),
        _m.get("minutos", "Minutos Requeridos"),
        _m.get("factor", "Factor Complejidad"),
        _m.get("minutos_ajustados", "Minutos Ajustados"),
        "Costo Minuto (Promedio)", 
        "Costo MO Total"
    ]
    
    row = 5
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row = 6  # Iniciar después de headers (row 5)
    for servicio, ecuaciones in ECUACIONES_SERVICIOS.items():
        # Buscar código del servicio
        codigo = f"SV{list(ECUACIONES_SERVICIOS.keys()).index(servicio) + 1:03d}"
        
        for grupo, minutos, factor in ecuaciones:
            ws[f'A{row}'] = codigo
            ws[f'B{row}'] = servicio
            ws[f'C{row}'] = grupo
            ws[f'D{row}'] = minutos
            ws[f'E{row}'] = factor
            ws[f'F{row}'] = f"=D{row}*E{row}"  # Minutos ajustados
            
            # Buscar costo por minuto del grupo en la Tabla Maestra de COSTO_POR_MINUTO
            # AHORA BUSCAMOS EN COLUMNAS G:H donde está la tabla correcta de promedios por minuto
            ws[f'G{row}'] = f"=IFERROR(VLOOKUP(C{row},COSTO_POR_MINUTO!$G:$H,2,FALSE),0)"
            
            # Costo Total MO de este recurso para este servicio
            ws[f'H{row}'] = f"=F{row}*G{row}"
            
            # Aplicar estilos
            aplicar_estilo_celda(ws[f'A{row}'], "calculo")
            aplicar_estilo_celda(ws[f'B{row}'], "calculo")
            aplicar_estilo_celda(ws[f'C{row}'], "calculo")
            aplicar_estilo_celda(ws[f'D{row}'], "input")
            aplicar_estilo_celda(ws[f'E{row}'], "input")
            aplicar_estilo_celda(ws[f'F{row}'], "resultado")
            aplicar_estilo_celda(ws[f'G{row}'], "calculo")
            aplicar_estilo_celda(ws[f'H{row}'], "resultado")
            
            # Formato
            ws[f'F{row}'].number_format = '#,##0.0'
            ws[f'G{row}'].number_format = '$#,##0'
            ws[f'H{row}'].number_format = '$#,##0'
            
            row += 1
    
    ajustar_columnas(ws, {
        'A': 15, 'B': 38, 'C': 30, 'D': 20, 'E': 20, 'F': 20, 'G': 20, 'H': 20
    })

