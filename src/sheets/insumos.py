"""
Generador de la hoja INSUMOS
"""
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas
from ..data.insumos_data import INSUMOS_POR_SERVICIO


def crear_hoja_insumos(wb):
    """Crea la hoja INSUMOS con costos de materiales"""
    ws = wb.create_sheet("INSUMOS")
    
    # Título
    ws['A1'] = "INSUMOS DIRECTOS POR SERVICIO"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A1:F1')
    
    # Encabezados dinámicos desde el mapeo
    _m = config._mapper.get_columnas_estandar("insumos")
    headers = [
        _m.get("codigo_servicio", "Código Servicio"),
        _m.get("nombre_servicio", "Nombre Servicio"),
        _m.get("tipo_insumo", "Tipo Insumo"),
        _m.get("cantidad", "Cantidad"),
        _m.get("costo_unitario", "Costo Unitario"),
        _m.get("costo_total", "Costo Total Insumo")
    ]
    
    row = 3
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row = 4
    for servicio, insumos in INSUMOS_POR_SERVICIO.items():
        # Buscar código del servicio
        idx = config.SERVICIOS.index(servicio) + 1 if servicio in config.SERVICIOS else 1
        codigo = f"SV{idx:03d}"
        
        for tipo_insumo, cantidad, costo_unitario in insumos:
            ws[f'A{row}'] = codigo
            ws[f'B{row}'] = servicio
            ws[f'C{row}'] = tipo_insumo
            ws[f'D{row}'] = cantidad
            ws[f'E{row}'] = costo_unitario
            ws[f'F{row}'] = f"=D{row}*E{row}"
            
            # Aplicar estilos
            aplicar_estilo_celda(ws[f'A{row}'], "calculo")
            aplicar_estilo_celda(ws[f'B{row}'], "calculo")
            aplicar_estilo_celda(ws[f'C{row}'], "calculo")
            aplicar_estilo_celda(ws[f'D{row}'], "input")
            aplicar_estilo_celda(ws[f'E{row}'], "input")
            aplicar_estilo_celda(ws[f'F{row}'], "resultado")
            
            # Formato
            ws[f'E{row}'].number_format = '$#,##0'
            ws[f'F{row}'].number_format = '$#,##0'
            
            row += 1
    
    ajustar_columnas(ws, {
        'A': 15, 'B': 38, 'C': 35, 'D': 12, 'E': 18, 'F': 20
    })
