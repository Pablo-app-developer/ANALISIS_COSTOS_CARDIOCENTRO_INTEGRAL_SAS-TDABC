"""
Generador de la hoja NOMINA
"""
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import random
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas


def crear_hoja_nomina(wb):
    """Crea la hoja NOMINA con estructura salarial"""
    ws = wb.create_sheet("NOMINA")
    
    # Título
    ws['A1'] = "ESTRUCTURA SALARIAL Y NÓMINA"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A1:H1')
    
    # Encabezados dinámicos desde el mapeo
    _m = config._mapper.get_columnas_estandar("nomina")
    headers = [
        _m.get("id", "ID"),
        _m.get("grupo", "Grupo Ocupacional"),
        _m.get("salario", "Salario Base Mensual"),
        _m.get("factor", "Factor Prestacional"),
        _m.get("costo_total", "Costo Total Mensual"),
        _m.get("cantidad", "Cantidad Personal"),
        _m.get("total_grupo", "Costo Total Grupo"),
        _m.get("centro", "Sede")
    ]
    
    row = 3
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = alignment
    
    # Datos de nómina
    row = 4
    id_counter = 1
    for grupo, salario in config.GRUPOS_OCUPACIONALES:
        for sede in config.SEDES:
            cantidad = random.randint(2, 8)
            
            ws[f'A{row}'] = f"G{id_counter:03d}"
            ws[f'B{row}'] = grupo
            ws[f'C{row}'] = salario
            ws[f'D{row}'] = "=PARAMETROS!$B$10"  # Factor prestacional sector salud (40.77%)
            ws[f'E{row}'] = f"=C{row}*D{row}"  # Costo total mensual
            ws[f'F{row}'] = cantidad
            ws[f'G{row}'] = f"=E{row}*F{row}"  # Costo total grupo
            ws[f'H{row}'] = sede
            
            # Aplicar estilos
            aplicar_estilo_celda(ws[f'A{row}'], "calculo")
            aplicar_estilo_celda(ws[f'B{row}'], "calculo")
            aplicar_estilo_celda(ws[f'C{row}'], "input")
            aplicar_estilo_celda(ws[f'D{row}'], "calculo")
            aplicar_estilo_celda(ws[f'E{row}'], "resultado")
            aplicar_estilo_celda(ws[f'F{row}'], "input")
            aplicar_estilo_celda(ws[f'G{row}'], "resultado")
            aplicar_estilo_celda(ws[f'H{row}'], "calculo")
            
            # Formato moneda
            ws[f'C{row}'].number_format = '$#,##0'
            ws[f'E{row}'].number_format = '$#,##0'
            ws[f'G{row}'].number_format = '$#,##0'
            
            row += 1
            id_counter += 1
    
    # Totales
    ws[f'B{row}'] = "TOTAL NÓMINA MENSUAL"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=SUM(G4:G{row-1})"
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'G{row}'].number_format = '$#,##0'
    font, fill, alignment, border = crear_estilo_header()
    ws[f'B{row}'].fill = fill
    ws[f'G{row}'].fill = fill
    ws.merge_cells(f'B{row}:F{row}')
    
    ajustar_columnas(ws, {
        'A': 10, 'B': 30, 'C': 20, 'D': 18, 
        'E': 20, 'F': 18, 'G': 20, 'H': 25
    })
