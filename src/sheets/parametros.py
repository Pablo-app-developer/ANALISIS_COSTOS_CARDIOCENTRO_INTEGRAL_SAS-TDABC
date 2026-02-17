"""
Generador de la hoja PARAMETROS
"""
from openpyxl.styles import Font, Alignment
from datetime import datetime
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas


def crear_hoja_parametros(wb):
    """Crea la hoja PARAMETROS con configuración general"""
    ws = wb.create_sheet("PARAMETROS")
    
    # Título
    ws['A1'] = config.NOMBRE_EMPRESA
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = "PARÁMETROS GENERALES DEL MODELO TDABC"
    ws['A2'].font = Font(name='Calibri', size=12, bold=True)
    ws.merge_cells('A2:D2')
    
    # Fecha de actualización
    ws['A3'] = f"Fecha de actualización: {datetime.now().strftime('%d/%m/%Y')}"
    ws.merge_cells('A3:D3')
    
    # Parámetros operacionales
    row = 5
    ws[f'A{row}'] = "PARÁMETROS OPERACIONALES"
    font, fill, alignment, border = crear_estilo_header()
    ws[f'A{row}'].font = font
    ws[f'A{row}'].fill = fill
    ws.merge_cells(f'A{row}:B{row}')
    
    parametros = [
        ("Horas mensuales por colaborador", config.CAPACIDAD_MENSUAL_MINUTOS / 60, "horas"),
        ("Días hábiles al mes", 22, "días"),
        ("Porcentaje capacidad práctica", 85, "%"),
        ("Factor prestacional sector salud", 1.4077, "factor"),
        ("Prima de servicios", 8.33, "%"),
        ("Cesantías", 8.33, "%"),
        ("Intereses cesantías", 1.00, "%"),
        ("Vacaciones", 4.17, "%"),
        ("ARL (Riesgo III - CATEROGI)", 2.436, "%"),
        ("Salud EPS (empleador)", 0, "%"),
        ("Pensión (empleador)", 12.5, "%"),
        ("Caja compensación", 4.0, "%"),
        ("ICBF (exonerado sector salud)", 0, "%"),
        ("SENA (exonerado sector salud)", 0, "%"),
        ("TOTAL PRESTACIONES LEGALES", 40.77, "%"),
        ("Inflación anual proyectada", 5.5, "%"),
        ("Tasa cambio USD (referencial)", 4100, "COP"),
    ]
    
    row += 1
    ws[f'A{row}'] = "Parámetro"
    ws[f'B{row}'] = "Valor"
    ws[f'C{row}'] = "Unidad"
    for col in ['A', 'B', 'C']:
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
    
    for param, valor, unidad in parametros:
        row += 1
        ws[f'A{row}'] = param
        ws[f'B{row}'] = valor
        ws[f'C{row}'] = unidad
        aplicar_estilo_celda(ws[f'A{row}'], "calculo")
        aplicar_estilo_celda(ws[f'B{row}'], "input")
        aplicar_estilo_celda(ws[f'C{row}'], "calculo")
    
    # Nota explicativa sobre prestaciones sector salud
    row += 2
    ws[f'A{row}'] = "NOTA: Empresas del sector salud en Colombia"
    ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color="2C3E50")
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws[f'A{row}'] = "Las empresas del sector salud están EXONERADAS del pago de parafiscales (ICBF 3% + SENA 2%)."
    ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color="2C3E50")
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws[f'A{row}'] = "Sin embargo, SÍ pagan EPS (8.5%), Pensión (12%), Caja Compensación (4%) y ARL según riesgo."
    ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True, color="2C3E50")
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f'A{row}:C{row}')
    
    # Parámetros de distribución de costos indirectos
    row += 3
    ws[f'A{row}'] = "DISTRIBUCIÓN DE COSTOS INDIRECTOS"
    font, fill, alignment, border = crear_estilo_header()
    ws[f'A{row}'].font = font
    ws[f'A{row}'].fill = fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Concepto"
    ws[f'B{row}'] = "Criterio"
    for col in ['A', 'B']:
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
    
    criterios = [
        ("Servicios generales", "Por minutos consumidos"),
        ("Administración", "Por volumen de producción"),
        ("Infraestructura", "Por área utilizada (m²)"),
    ]
    
    for concepto, criterio in criterios:
        row += 1
        ws[f'A{row}'] = concepto
        ws[f'B{row}'] = criterio
        aplicar_estilo_celda(ws[f'A{row}'], "calculo")
        aplicar_estilo_celda(ws[f'B{row}'], "calculo")
    
    # Información de sedes
    row += 3
    ws[f'A{row}'] = "SEDES OPERATIVAS"
    font, fill, alignment, border = crear_estilo_header()
    ws[f'A{row}'].font = font
    ws[f'A{row}'].fill = fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers = ["Código", "Nombre Sede", "Ciudad", "Área (m²)"]
    from openpyxl.utils import get_column_letter
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
    
    sedes_info = [
        ("S01", "Sede Norte - Bogotá", "Bogotá", 450),
        ("S02", "Sede Sur - Medellín", "Medellín", 380),
        ("S03", "Sede Centro - Cali", "Cali", 420),
    ]
    
    for codigo, nombre, ciudad, area in sedes_info:
        row += 1
        ws[f'A{row}'] = codigo
        ws[f'B{row}'] = nombre
        ws[f'C{row}'] = ciudad
        ws[f'D{row}'] = area
        for col in ['A', 'B', 'C']:
            aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
        aplicar_estilo_celda(ws[f'D{row}'], "input")
    
    ajustar_columnas(ws, {'A': 35, 'B': 30, 'C': 20, 'D': 15})
