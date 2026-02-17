"""
Generador de la hoja PRODUCCION
"""
from openpyxl.styles import Font, Alignment
from pathlib import Path
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas
from ..utils import crear_tabla
from ..generators.produccion_generator import generar_datos_produccion


def crear_hoja_produccion(wb, datos_reales_path: str = None):
    """
    Crea la hoja PRODUCCION con volúmenes y facturación.
    
    Args:
        wb: Workbook de openpyxl
        datos_reales_path: Ruta opcional a archivo Excel/CSV con datos reales de producción.
                          Si no se proporciona, genera datos simulados.
    """
    """Crea la hoja PRODUCCION con volúmenes y facturación"""
    ws = wb.create_sheet("PRODUCCION")

    # Título
    ws['A1'] = "PRODUCCIÓN MENSUAL Y FACTURACIÓN"
    ws['A1'].font = Font(name=config.FUENTE_BASE, size=14, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A1:G1')

    # Encabezados dinámicos desde el mapeo
    _m = config._mapper.get_columnas_estandar("produccion")
    headers = [
        _m.get("codigo", "Código"),
        _m.get("servicio", "Servicio"),
        _m.get("centro", "Sede"),
        _m.get("cliente", "Aseguradora"),
        _m.get("cantidad", "Cantidad"),
        _m.get("valor_unitario", "Valor Facturado Promedio"),
        _m.get("valor_total", "Total Facturado")
    ]

    row = 3
    for i, header in enumerate(headers):
        col = chr(65 + i)
        ws[f'{col}{row}'] = header
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    # Determinar fuente de datos: reales o simulados
    if datos_reales_path and Path(datos_reales_path).exists():
        # Importar datos reales
        from ..importador_produccion import ImportadorProduccion
        
        importador = ImportadorProduccion()
        
        # Detectar tipo de archivo
        if datos_reales_path.endswith('.csv'):
            exito = importador.cargar_desde_csv(datos_reales_path)
        else:
            exito = importador.cargar_desde_excel(datos_reales_path)
        
        if exito:
            datos_produccion = importador.obtener_datos_produccion()
            print(f"[INFO] Datos reales importados desde: {datos_reales_path}")
            print(importador.obtener_reporte())
        else:
            print(f"[ERROR] No se pudieron importar datos reales:")
            print(importador.obtener_reporte())
            print("[WARN] Generando datos simulados como fallback...")
            datos_produccion = _generar_datos_simulados()
    else:
        # Generar datos simulados
        if datos_reales_path:
            print(f"[WARN] Archivo no encontrado: {datos_reales_path}")
            print("[INFO] Generando datos simulados...")
        datos_produccion = _generar_datos_simulados()

    # Presentación: escribir datos en Excel
    row = 4
    for dato in datos_produccion:
        ws[f'A{row}'] = dato['codigo']
        ws[f'B{row}'] = dato['servicio']
        ws[f'C{row}'] = dato['sede']
        ws[f'D{row}'] = dato['aseguradora']
        ws[f'E{row}'] = dato['cantidad']
        ws[f'F{row}'] = dato['valor_unitario']
        ws[f'G{row}'] = f"=E{row}*F{row}"

        # Aplicar estilos
        aplicar_estilo_celda(ws[f'A{row}'], "calculo")
        aplicar_estilo_celda(ws[f'B{row}'], "calculo")
        aplicar_estilo_celda(ws[f'C{row}'], "calculo")
        aplicar_estilo_celda(ws[f'D{row}'], "calculo")
        aplicar_estilo_celda(ws[f'E{row}'], "input")
        aplicar_estilo_celda(ws[f'F{row}'], "input")
        aplicar_estilo_celda(ws[f'G{row}'], "resultado")

        # Formato
        ws[f'F{row}'].number_format = '$#,##0'
        ws[f'G{row}'].number_format = '$#,##0'

        row += 1

    # Crear Tabla Oficial
    crear_tabla(ws, "TablaProduccion", f"A3:G{row-1}")

    # Totales
    ws[f'B{row}'] = "TOTAL FACTURACIÓN MENSUAL"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f"=SUM(E4:E{row-1})"
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=SUM(G4:G{row-1})"
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'G{row}'].number_format = '$#,##0'
    font, fill, alignment, border = crear_estilo_header()
    ws[f'B{row}'].fill = fill
    ws[f'E{row}'].fill = fill
    ws[f'G{row}'].fill = fill
    ws.merge_cells(f'B{row}:D{row}')

    ajustar_columnas(ws, {
        'A': 12, 'B': 35, 'C': 22, 'D': 20, 'E': 12, 'F': 18, 'G': 18
    })


def _generar_datos_simulados():
    """
    Genera datos simulados de producción para demostración.
    
    Returns:
        Lista de diccionarios con datos de producción simulados
    """
    servicios_completos = config._mapper.get_servicios_completos() or []
    categorias_info = config._mapper.get_categorias_info() or {}
    servicios_dict = {s["nombre"]: s for s in servicios_completos}
    
    return generar_datos_produccion(
        config.SERVICIOS,
        servicios_dict,
        categorias_info,
        config.SEDES,
        config.ASEGURADORAS
    )
