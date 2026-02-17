"""
Generador de la hoja RESUMEN_EJECUTIVO
"""
from openpyxl.styles import Font, Alignment
from datetime import datetime
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas


def crear_hoja_resumen_ejecutivo(wb):
    """Crea la hoja RESUMEN_EJECUTIVO con análisis consolidado y Conciliación"""
    ws = wb.create_sheet("RESUMEN_EJECUTIVO")
    
    # Título principal
    ws['A1'] = config.NOMBRE_EMPRESA
    ws['A1'].font = Font(name=config.FUENTE_BASE, size=18, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "RESUMEN EJECUTIVO - ANÁLISIS DE RENTABILIDAD (Enero)"
    ws['A2'].font = Font(name=config.FUENTE_BASE, size=14, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A2:F2')
    
    ws['A3'] = f"Período: {datetime.now().strftime('%B %Y')}"
    ws['A3'].font = Font(name=config.FUENTE_BASE, size=11, italic=True)
    ws.merge_cells('A3:F3')
    
    # SECCIÓN 1: Indicadores Globales
    row = 5
    ws[f'A{row}'] = "INDICADORES GLOBALES DE RESULTADOS"
    font, fill, alignment, border = crear_estilo_header()
    ws[f'A{row}'].font = font
    ws[f'A{row}'].fill = fill
    ws.merge_cells(f'A{row}:B{row}')
    
    # Obtener mapeos para fórmulas
    m_p = config._mapper.get_columnas_estandar("produccion")
    m_c = config._mapper.get_columnas_estandar("costeo_servicios")
    m_i = config._mapper.get_columnas_estandar("contabilidad")

    indicadores = [
        ("Ingresos Operacionales (Facturación)", f"=SUM(TablaProduccion[{m_p.get('valor_total', 'Total Facturado')}])", '$#,##0'),
        ("Total Costos Asignados (TDABC)", f"=SUMPRODUCT(TablaCosteo[{m_c.get('total', 'Costo Unitario Total')}],TablaCosteo[{m_c.get('volumen', 'Volumen Mes')}])", '$#,##0'),
        (f"   - {m_c.get('mo', 'Costo Personal Directo')}", f"=SUMPRODUCT(TablaCosteo[{m_c.get('mo', 'Costo MO Directa')}],TablaCosteo[{m_c.get('volumen', 'Volumen Mes')}])", '$#,##0'),
        (f"   - {m_c.get('insumos', 'Costo Insumos')}", f"=SUMPRODUCT(TablaCosteo[{m_c.get('insumos', 'Costo Insumos')}],TablaCosteo[{m_c.get('volumen', 'Volumen Mes')}])", '$#,##0'),
        (f"   - {m_c.get('cif', 'Costos Indirectos (CIF)')}", f"=SUMPRODUCT(TablaCosteo[{m_c.get('cif', 'Costo CIF (Indirecto)')}],TablaCosteo[{m_c.get('volumen', 'Volumen Mes')}])", '$#,##0'),
        ("", "", ""),  # Línea en blanco
        ("UTILIDAD OPERACIONAL", "=B7-B8", '$#,##0'),
        ("MARGEN OPERACIONAL %", "=IF(B7>0,B13/B7,0)", '0.0%'),
        ("", "", ""),  # Línea en blanco
        ("Total Servicios Prestados", f"=SUM(TablaProduccion[{m_p.get('cantidad', 'Cantidad')}])", '#,##0'),
        ("Precio Promedio por Servicio", "=IF(B16>0,B7/B16,0)", '$#,##0'),
        ("Costo Promedio por Servicio", "=IF(B16>0,B8/B16,0)", '$#,##0'),
    ]

    
    row += 1
    ws[f'A{row}'] = "Indicador"
    ws[f'B{row}'] = "Valor"
    ws[f'C{row}'] = ""
    for col in ['A', 'B']:
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
    
    for indicador, formula, formato in indicadores:
        row += 1
        ws[f'A{row}'] = indicador
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = formato
        
        if "MARGEN" in indicador:
            aplicar_estilo_celda(ws[f'A{row}'], "resultado")
            aplicar_estilo_celda(ws[f'B{row}'], "resultado")
        else:
            aplicar_estilo_celda(ws[f'A{row}'], "calculo")
            aplicar_estilo_celda(ws[f'B{row}'], "calculo")
        
        # Sangría visual
        if indicador.startswith("   -"):
             ws[f'A{row}'].alignment = Alignment(indent=2)

    # SECCIÓN 1.5: CONCILIACIÓN DE COSTOS
    row += 3
    ws[f'A{row}'] = "CONCILIACIÓN DE COSTOS (CONTABLE VS DISTRIBUIDO)"
    font, fill, alignment, border = crear_estilo_header()
    ws[f'A{row}'].font = font
    ws[f'A{row}'].fill = fill
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws[f'A{row}'] = f"Concepto (Fuente: COSTOS_INDIRECTOS)"
    ws[f'B{row}'] = "Costo Contable (Gastado)"
    ws[f'C{row}'] = "Costo Asignado (TDABC)"
    
    for col in ['A', 'B', 'C']:
         ws[f'{col}{row}'].font = Font(name=config.FUENTE_BASE, size=10, bold=True)
         ws[f'{col}{row}'].border = border
         
    # Cuentas dinámicas
    cuenta_mp = config._mapper.get_cuenta_materia_prima().get("codigo", "7105")[:2] + "*"
    cuenta_mo = config._mapper.get_cuenta_mano_obra().get("codigo", "7205")[:2] + "*"

    row += 1
    ws[f'A{row}'] = "Materia Prima / Insumos (71)"
    ws[f'B{row}'] = f'=SUMIFS(TablaIndirectos[{m_i.get("valor", "Valor Mensual")}],TablaIndirectos[{m_i.get("cuenta", "Código Cuenta")}],"71*")' 
    ws[f'C{row}'] = f"=SUMPRODUCT(TablaCosteo[{m_c.get('insumos', 'Costo Insumos')}],TablaCosteo[{m_c.get('volumen', 'Volumen Mes')}])"
    
    row += 1
    ws[f'A{row}'] = "Mano de Obra Directa (72)"
    ws[f'B{row}'] = f'=SUMIFS(TablaIndirectos[{m_i.get("valor", "Valor Mensual")}],TablaIndirectos[{m_i.get("cuenta", "Código Cuenta")}],"72*")'
    ws[f'C{row}'] = f"=SUMPRODUCT(TablaCosteo[{m_c.get('mo', 'Costo MO Directa')}],TablaCosteo[{m_c.get('volumen', 'Volumen Mes')}])"
    
    row += 1
    ws[f'A{row}'] = "Costos Indirectos CIF (73)"
    ws[f'B{row}'] = f'=SUMIFS(TablaIndirectos[{m_i.get("valor", "Valor Mensual")}],TablaIndirectos[{m_i.get("cuenta", "Código Cuenta")}],"73*")'
    ws[f'C{row}'] = f"=SUMPRODUCT(TablaCosteo[{m_c.get('cif', 'Costo CIF (Indirecto)')}],TablaCosteo[{m_c.get('volumen', 'Volumen Mes')}])"
    
    # Aplicar formatos
    for r in range(row-2, row+1):
        for col in ['B', 'C']:
            ws[f'{col}{r}'].number_format = '$#,##0'
            aplicar_estilo_celda(ws[f'{col}{r}'], "calculo")
        aplicar_estilo_celda(ws[f'A{r}'])

    row += 1
    ws[f'A{row}'] = "DIFERENCIA (CAPACIDAD OCIOSA)"
    ws[f'A{row}'].font = Font(name=config.FUENTE_BASE, bold=True, color="FF0000")
    
    r_start = row-3
    r_end = row-1
    ws[f'B{row}'] = f"=SUM(B{r_start}:B{r_end})-SUM(C{r_start}:C{r_end})"
    
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'].alignment = Alignment(horizontal="center")
    ws[f'B{row}'].number_format = '$#,##0'
    
    # Agregar % de utilización de capacidad
    row += 1
    ws[f'A{row}'] = "% UTILIZACIÓN DE CAPACIDAD"
    ws[f'A{row}'].font = Font(name=config.FUENTE_BASE, bold=True, color="0070C0")
    ws[f'B{row}'] = f"=IF(SUM(B{r_start}:B{r_end})>0,SUM(C{r_start}:C{r_end})/SUM(B{r_start}:B{r_end}),0)"
    ws[f'B{row}'].number_format = '0.0%'
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'].alignment = Alignment(horizontal="center")
    aplicar_estilo_celda(ws[f'B{row}'], "resultado")
    
    row += 1
    ws[f'A{row}'] = "NOTA IMPORTANTE SOBRE CAPACIDAD OCIOSA:"
    ws[f'A{row}'].font = Font(name=config.FUENTE_BASE, bold=True, color="CC0000")
    
    row += 1
    nota = (
        "1. Diferencia en '72' (Mano de Obra): CAPACIDAD OCIOSA DE PERSONAL. "
        "Indica tiempo pagado a empleados que NO se usó en atención a pacientes.\n"
        "2. Diferencia en '73' (Ind. Fabricación): CAPACIDAD OCIOSA DE INFRAESTRUCTURA. "
        "Indica costos fijos de planta (arriendo, depreciación) no absorbidos por falta de volumen.\n"
        "3. Valores Positivos (+) = Pérdida por Capacidad Ociosa (Gasto > Uso).\n"
        "4. Valores Negativos (-) = Sobre-ejecución / Eficiencia superior a la estándar."
    )
    ws[f'A{row}'] = nota
    ws[f'A{row}'].font = Font(name=config.FUENTE_BASE, italic=True, size=9)
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 75
    ws.merge_cells(f'A{row}:F{row}')

    # SECCIÓN 2: RENTABILIDAD POR SERVICIO
    row += 3
    ws[f'A{row}'] = "RENTABILIDAD POR SERVICIO (TODOS)"
    font, fill, alignment, border = crear_estilo_header()
    ws[f'A{row}'].font = font
    ws[f'A{row}'].fill = fill
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = "Servicio"
    ws[f'B{row}'] = "Volumen"
    ws[f'C{row}'] = "Costo Total"
    ws[f'D{row}'] = "Facturación Total"
    ws[f'E{row}'] = "Margen Total"
    ws[f'F{row}'] = "Margen %"
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
    
    # Listar TODOS los servicios
    for servicio in config.SERVICIOS:
        row += 1
        ws[f'A{row}'] = servicio
        ws[f'B{row}'] = f"=SUMIFS(TablaProduccion[{m_p.get('cantidad', 'Cantidad')}],TablaProduccion[{m_p.get('servicio', 'Servicio')}],A{row})"
        ws[f'C{row}'] = f"=SUMPRODUCT((TablaCosteo[{m_c.get('servicio', 'Servicio')}]=A{row})*(TablaCosteo[{m_c.get('total', 'Costo Unitario Total')}])*(TablaCosteo[{m_c.get('volumen', 'Volumen Mes')}]))"
        ws[f'D{row}'] = f"=SUMIFS(TablaProduccion[{m_p.get('valor_total', 'Total Facturado')}],TablaProduccion[{m_p.get('servicio', 'Servicio')}],A{row})"
        ws[f'E{row}'] = f"=D{row}-C{row}"
        ws[f'F{row}'] = f"=IF(D{row}>0,E{row}/D{row},0)"
        
        for col in ['A']:
            aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
        for col in ['B', 'C', 'D', 'E', 'F']:
            aplicar_estilo_celda(ws[f'{col}{row}'], "resultado")
        
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'].number_format = '$#,##0'
        ws[f'E{row}'].number_format = '$#,##0'
        ws[f'F{row}'].number_format = '0.0%'
    
    # SECCIÓN 3: Análisis por Sede
    row += 3
    ws[f'A{row}'] = "ANÁLISIS POR SEDE"
    font, fill, alignment, border = crear_estilo_header()
    ws[f'A{row}'].font = font
    ws[f'A{row}'].fill = fill
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = m_p.get("centro", "Sede")
    ws[f'B{row}'] = "Facturación"
    ws[f'C{row}'] = "Costos"
    ws[f'D{row}'] = "Margen"
    ws[f'E{row}'] = "Margen %"
    for col in ['A', 'B', 'C', 'D', 'E']:
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
    
    for sede in config.SEDES:
        row += 1
        ws[f'A{row}'] = sede
        ws[f'B{row}'] = f'=SUMIFS(TablaProduccion[{m_p.get("valor_total", "Total Facturado")}],TablaProduccion[{m_p.get("centro", "Sede")}],A{row})'
        ws[f'C{row}'] = f'=SUMPRODUCT((TablaCosteo[{m_c.get("centro", "Sede")}]=A{row})*(TablaCosteo[{m_c.get("total", "Costo Unitario Total")}])*(TablaCosteo[{m_c.get("volumen", "Volumen Mes")}]))'
        ws[f'D{row}'] = f'=B{row}-C{row}'
        ws[f'E{row}'] = f'=IF(B{row}>0,D{row}/B{row},0)'
        
        for col in ['A']:
            aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
        for col in ['B', 'C', 'D', 'E']:
            aplicar_estilo_celda(ws[f'{col}{row}'], "resultado")
        
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'].number_format = '$#,##0'
        ws[f'E{row}'].number_format = '0.0%'
    
    # SECCIÓN 4: Notas importantes
    row += 3
    ws[f'A{row}'] = "NOTAS ADICIONALES:"
    ws[f'A{row}'].font = Font(name=config.FUENTE_BASE, bold=True)
    ws[f'A{row}'].alignment = Alignment(vertical='top')
    ws.merge_cells(f'A{row}:F{row+3}')
    ws[f'A{row}'] = "1. El modelo utiliza referencias directas para cálculo de materiales, asegurando integridad referencial.\n2. La Tabla 'COSTOS_INDIRECTOS' centraliza toda la contabilidad (Insumos 71 + MO 72 + Indirectos 73).\n3. La Capacidad Ociosa refleja la diferencia entre los recursos pagados y los consumidos."
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                     
    ajustar_columnas(ws, {
        'A': 45, 'B': 20, 'C': 20, 'D': 20, 'E': 15, 'F': 15
    })
