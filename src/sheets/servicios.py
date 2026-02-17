"""
Generador de la hoja SERVICIOS
"""
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from .. import config
from ..styles import crear_estilo_header, aplicar_estilo_celda, ajustar_columnas
from ..generators.servicios_generator import generar_datos_servicios


def crear_hoja_servicios(wb):
    """Crea la hoja SERVICIOS con catálogo de servicios"""
    ws = wb.create_sheet("SERVICIOS")
    
    # Título
    ws['A1'] = "CATÁLOGO DE SERVICIOS"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=config.COLOR_HEADER)
    ws.merge_cells('A1:F1')
    
    # Encabezados dinámicos desde el mapeo
    _m = config._mapper.get_columnas_estandar("servicios")
    headers = [
        _m.get("codigo", "Código"),
        _m.get("nombre", "Nombre del Servicio"),
        _m.get("categoria", "Categoría"),
        _m.get("complejidad", "Complejidad"),
        _m.get("insumos", "Requiere Insumos"),
        _m.get("estado", "Estado")
    ]
    
    row = 3
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{row}'] = header
        font, fill, alignment, border = crear_estilo_header()
        ws[f'{col}{row}'].font = font
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Generar datos usando el generador (lógica de negocio separada)
    servicios_completos = config._mapper.get_servicios_completos() or []
    categorias_info = config._mapper.get_categorias_info() or {}
    servicios_data = generar_datos_servicios(servicios_completos, categorias_info)
    
    # Presentación: escribir datos en Excel
    row = 4
    for servicio in servicios_data:
        ws[f'A{row}'] = servicio['codigo']
        ws[f'B{row}'] = servicio['nombre']
        ws[f'C{row}'] = servicio['categoria']
        ws[f'D{row}'] = servicio['complejidad']
        ws[f'E{row}'] = "Sí" if servicio['requiere_insumos'] else "No"
        ws[f'F{row}'] = servicio['estado']
        
        # Aplicar estilos
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            aplicar_estilo_celda(ws[f'{col}{row}'], "calculo")
        
        row += 1
    
    ajustar_columnas(ws, {
        'A': 12, 'B': 38, 'C': 25, 'D': 15, 'E': 18, 'F': 12
    })
