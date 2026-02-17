"""
Utilidades de estilos para Excel
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from . import config


def crear_estilo_header():
    """Crea el estilo para encabezados"""
    font = Font(name=config.FUENTE_BASE, size=11, bold=True, color="FFFFFF")
    fill = PatternFill(start_color=config.COLOR_HEADER, end_color=config.COLOR_HEADER, fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return font, fill, alignment, border


def aplicar_estilo_celda(cell, tipo="normal"):
    """Aplica estilo a una celda según su tipo"""
    if tipo == "input":
        cell.fill = PatternFill(start_color=config.COLOR_INPUT, end_color=config.COLOR_INPUT, fill_type="solid")
    elif tipo == "calculo":
        cell.fill = PatternFill(start_color=config.COLOR_CALCULO, end_color=config.COLOR_CALCULO, fill_type="solid")
    elif tipo == "resultado":
        cell.fill = PatternFill(start_color=config.COLOR_RESULTADO, end_color=config.COLOR_RESULTADO, fill_type="solid")


def ajustar_columnas(ws, columnas_info):
    """Ajusta el ancho de las columnas"""
    for col, width in columnas_info.items():
        ws.column_dimensions[col].width = width
